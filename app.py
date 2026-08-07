import streamlit as st
import pandas as pd
import numpy as np
import re
import ast
from io import BytesIO
from PIL import Image
import pdfplumber
import os
from datetime import datetime, timedelta
import html as html_lib
import streamlit.components.v1 as components

APP_VERSION = "SIGMA FINANCIAL REPORTS - PDF SAFE V2 + REPORT VIEW - 07.08.2026"


def clean_number(x):
    """
    Универзално чистење на бројки за локално и online.
    Поддржува:
    - MK/EU формат: 1.234,56
    - US формат: 1,234.56
    - цели броеви: 1234
    - негативни во заграда: (1.234,56)
    """
    if pd.isna(x):
        return 0.0

    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x).strip()
    s = s.replace("\xa0", "").replace(" ", "")

    if s.lower() in ["", "nan", "none", "-", "—"]:
        return 0.0

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]

    if s.startswith("-"):
        negative = True
        s = s[1:]

    # Ако има и точка и запирка, последниот сепаратор е децимален
    if "." in s and "," in s:
        if s.rfind(",") > s.rfind("."):
            # 1.234,56 -> 1234.56
            s = s.replace(".", "").replace(",", ".")
        else:
            # 1,234.56 -> 1234.56
            s = s.replace(",", "")

    elif "," in s:
        # 1,234 или 12,345,678 = илјадарски US формат
        if re.fullmatch(r"\d{1,3}(,\d{3})+", s):
            s = s.replace(",", "")
        else:
            # 1234,56 = децимална запирка
            s = s.replace(",", ".")

    elif "." in s:
        # 1.234 или 12.345.678 = илјадарски MK/EU формат
        if re.fullmatch(r"\d{1,3}(\.\d{3})+", s):
            s = s.replace(".", "")
        # инаку 1234.56 останува како децимална точка

    try:
        value = float(s)
        return -value if negative else value
    except Exception:
        return 0.0


def normalize_dataframe_numbers(df):
    """
    Оваа функција ги чисти сите колони што личат на износи.
    Не ги допира колони како конто, опис, назив, АОП.
    """
    # Работиме на независна копија за да избегнеме pandas chained-assignment
    # warnings и идни Copy-on-Write несогласувања.
    df = df.copy()

    skip_words = [
        "konto", "конто",
        "opis", "опис",
        "naziv", "назив",
        "aop", "аоп",
        "sifra", "шифра",
        "ime", "име"
    ]

    for col in df.columns:
        col_lower = str(col).lower()

        if any(word in col_lower for word in skip_words):
            continue

        df.loc[:, col] = df[col].apply(clean_number)

    return df


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

pravila_file_path = os.path.join(BASE_DIR, "data", "pravila bilansi.xlsx")


from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

pdfmetrics.registerFont(
    TTFont("PDF_Font", "arial.ttf")
)

pdfmetrics.registerFont(
    TTFont("PDF_Font_Bold", "arialbd.ttf")
)

def calculate_formula(formula, values_dict):

    if pd.isna(formula):
        return 0

    formula = str(formula)

    parts = formula.split("+")

    total = 0

    for part in parts:

        part = part.strip()

        if part in values_dict:
            total += values_dict[part]

    return total


st.set_page_config(page_title="AOP Finansiski Izvestai", layout="wide")

# =========================
# PREMIUM SIDEBAR DESIGN
# =========================
st.markdown("""
<style>

/* Main sidebar background */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0f172a 0%, #111827 48%, #020617 100%);
    border-right: 1px solid rgba(255,255,255,0.08);
}

/* Sidebar inner spacing */
[data-testid="stSidebar"] > div:first-child {
    padding-top: 1.1rem;
}

/* Sidebar text default */
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span {
    color: #e5e7eb;
}

/* Sidebar input readability fix */
[data-testid="stSidebar"] input,
[data-testid="stSidebar"] input[type="text"],
[data-testid="stSidebar"] input[type="number"],
[data-testid="stSidebar"] textarea {
    color: #111827 !important;
    -webkit-text-fill-color: #111827 !important;
    caret-color: #111827 !important;
    background: #ffffff !important;
}

[data-testid="stSidebar"] input::placeholder {
    color: #6b7280 !important;
    -webkit-text-fill-color: #6b7280 !important;
}

/* Premium brand card */
.sigma-brand-card {
    background: linear-gradient(135deg, rgba(37,99,235,0.95), rgba(14,165,233,0.78));
    border: 1px solid rgba(255,255,255,0.18);
    border-radius: 18px;
    padding: 16px 14px;
    margin: 4px 0 14px 0;
    box-shadow: 0 10px 26px rgba(0,0,0,0.32);
}

.sigma-brand-title {
    font-size: 22px;
    font-weight: 800;
    letter-spacing: 0.3px;
    color: #ffffff;
    margin-bottom: 3px;
}

.sigma-brand-subtitle {
    font-size: 12px;
    color: rgba(255,255,255,0.86);
    line-height: 1.35;
}

.sigma-version-pill {
    display: inline-block;
    margin-top: 10px;
    padding: 5px 9px;
    border-radius: 999px;
    background: rgba(15,23,42,0.36);
    color: #ffffff;
    font-size: 11px;
    font-weight: 700;
}

/* Sidebar section labels */
.sigma-sidebar-section {
    margin: 16px 0 8px 0;
    color: #93c5fd;
    font-size: 12px;
    font-weight: 800;
    text-transform: uppercase;
    letter-spacing: 0.08em;
}

/* Radio block spacing */
[data-testid="stSidebar"] .stRadio > div {
    gap: 7px;
}

/* Radio option cards */
[data-testid="stSidebar"] .stRadio label {
    background: rgba(255,255,255,0.055);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 13px;
    padding: 10px 12px;
    margin-bottom: 4px;
    transition: all 0.18s ease;
}

[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(37,99,235,0.20);
    border-color: rgba(96,165,250,0.45);
    transform: translateX(3px);
}

/* Hide default tiny radio circle a bit */
[data-testid="stSidebar"] .stRadio label > div:first-child {
    opacity: 0.72;
}

/* Inputs inside sidebar */
[data-testid="stSidebar"] input {
    border-radius: 10px !important;
}

/* Sidebar buttons */
[data-testid="stSidebar"] button {
    border-radius: 12px !important;
    font-weight: 700 !important;
}

/* Expander styling */
[data-testid="stSidebar"] details {
    background: rgba(255,255,255,0.055);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 14px;
    padding: 4px 8px;
}

/* Small footer */
.sigma-sidebar-footer {
    margin-top: 18px;
    padding: 12px;
    border-radius: 14px;
    background: rgba(255,255,255,0.055);
    border: 1px solid rgba(255,255,255,0.08);
    font-size: 12px;
    line-height: 1.4;
    color: #cbd5e1;
}
            
/* TextInput */
.stTextInput input {
    color: #111827 !important;
    background-color: white !important;
    -webkit-text-fill-color: #111827 !important;
}

/* NumberInput */
.stNumberInput input {
    color: #111827 !important;
    background-color: white !important;
    -webkit-text-fill-color: #111827 !important;
}

/* Sidebar specific */
[data-testid="stSidebar"] .stTextInput input,
[data-testid="stSidebar"] .stNumberInput input {
    color: #111827 !important;
    background-color: white !important;
    -webkit-text-fill-color: #111827 !important;
}            

</style>
""", unsafe_allow_html=True)

st.sidebar.markdown(
    f"""
    <div class="sigma-brand-card">
        <div class="sigma-brand-title">📊 SIGMA</div>
        <div class="sigma-brand-subtitle">Accounting reporting dashboard<br>Финансиски извештаи и анализа</div>
        <div class="sigma-version-pill">{APP_VERSION}</div>
    </div>
    """,
    unsafe_allow_html=True
)

with st.sidebar.expander("🔎 System info", expanded=False):
    st.write("Pandas:", pd.__version__)
    st.write("NumPy:", np.__version__)
    st.write("Работна папка:", os.getcwd())

if st.sidebar.button("♻️ Освежи пресметки / исчисти cache"):
    try:
        st.cache_data.clear()
    except Exception:
        pass
    st.rerun()
USERS = {
    "sigma": "12345",
    "client1": "test123"
}




def premium_lock(module_name):
    st.warning(f"{module_name} е достапен само во PREMIUM верзија")
    st.info("Овој модул е отклучен само за PREMIUM корисници.")


st.markdown("""
<style>

/* TAB BAR */
.stTabs [data-baseweb="tab-list"] {
    gap: 12px;
    background-color: #f5f7fb;
    padding: 10px;
    border-radius: 14px;
}

/* TAB */
.stTabs [data-baseweb="tab"] {
    height: 50px;
    background-color: white;
    border-radius: 12px;
    padding: 10px 22px;
    color: #1f2937;
    font-weight: 600;
    border: 1px solid #e5e7eb;
    transition: all 0.25s ease;
}

/* HOVER */
.stTabs [data-baseweb="tab"]:hover {
    background-color: #eef2ff;
    color: #2563eb;
    transform: translateY(-2px);
}

/* ACTIVE TAB */
.stTabs [aria-selected="true"] {
    background: linear-gradient(90deg, #2563eb, #1d4ed8);
    color: white !important;
    box-shadow: 0 4px 12px rgba(37,99,235,0.35);
}

/* REMOVE RED LINE */
.stTabs [data-baseweb="tab-highlight"] {
    display: none;
}

</style>
""", unsafe_allow_html=True)


def login():
    st.sidebar.title("🔐 Login")

    username = st.sidebar.text_input("Корисник")
    password = st.sidebar.text_input("Лозинка", type="password")

    if st.sidebar.button("Најави се"):

        if username in USERS and USERS[username] == password:
            st.session_state["logged_in"] = True
            st.session_state["username"] = username
            st.rerun()
        else:
            st.sidebar.error("Погрешно корисничко име или лозинка")

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login()
    st.stop()

username = st.session_state["username"]

USERS_INFO = {
    "sigma": {
        "plan": "PREMIUM",
        "trial_start": None,
        "trial_days": None
    },
    "client1": {
        "plan": "TRIAL",
        "trial_start": "2026-05-28",
        "trial_days": 14
    }
}

user_info = USERS_INFO.get(username, {
    "plan": "FREE",
    "trial_start": None,
    "trial_days": None
})

user_plan = user_info["plan"]

# =========================
# TRIAL LOGIC
# =========================

if user_plan == "TRIAL":

    trial_start = datetime.strptime(
        user_info["trial_start"],
        "%Y-%m-%d"
    )

    trial_end = trial_start + timedelta(
        days=user_info["trial_days"]
    )

    days_left = (trial_end - datetime.now()).days

    if datetime.now() <= trial_end:

        user_plan = "PREMIUM"

        st.sidebar.info(
            f"""
🎁 TRIAL VERSION

Premium модулите се активни.

⏳ Преостанати денови: {days_left}

📅 Активно до:
{trial_end.strftime('%d.%m.%Y')}
"""
        )

    else:

        user_plan = "FREE"

        st.sidebar.warning(
            "🔒 TRIAL периодот е истечен"
        )

if os.path.exists("assets/sigma_logo.png"):
    logo = Image.open("assets/sigma_logo.png")
else:
    logo = Image.open("sigma_logo.png")


st.image(logo,width=420)
st.markdown("""
            ### Добредојдовте во SIGMA Accounting!
            Контактирајте не за повеќе информации и прилагодени решенија за вашата компанија:
            - 📧 Email:sigmaakaunting@gmail.com
            - 📞 Телефон: 078/229-057   
            - 🏠 Адреса: ул. 121 3-1 Тетово
            - 🌐 https://sigma-financial-reports.onrender.com
            🌏 SIGMA Accounting 
<style>
[data-testid="metric-container"] {
    background-color: #f8f9fa;
    border: 1px solid #e6e6e6;
    padding: 20px;
    border-radius: 14px;
}
[data-testid="metric-container"] label {
    font-size: 22px !important;
    font-weight: 700 !important;
}
[data-testid="metric-container"] p {
    font-size: 42px !important;
    font-weight: 800 !important;
}
</style>
""", unsafe_allow_html=True)

st.title("📊 SIGMA Accounting ")
st.caption("Автоматизирана финансиска анализа од заклучен лист")

zaklucen_file = st.file_uploader(
    "📤 Прикачи заклучен лист",
    type=["xls", "xlsx", "pdf"]
)

# =====================================================
# REPORT NAVIGATION
# Изборот на извештај се прави ПРЕД рендерирање на главната содржина.
# Така на екранот се прикажува само избраниот извештај, без потреба
# прво да се скрола низ заклучниот лист и дијагностиките.
# =====================================================
REPORT_OPTIONS = [
    "🏠 Заклучен лист",
    "📈 Биланс на успех",
    "🏦 Биланс на состојба",
    "💧 Cash Flow",
    "📊 KPI Dashboard",
    "🧾 Сеопфатна добивка",
    "💎 Valuation Dashboard",
    "📍 Break-even Analysis",
    "📄 CFO Snapshot",
    "📘 CFO Извештај",
]

if zaklucen_file is not None:
    st.sidebar.markdown(
        '<div class="sigma-sidebar-section">Навигација</div>',
        unsafe_allow_html=True
    )
    izbran_izvestaj = st.sidebar.radio(
        "",
        REPORT_OPTIONS,
        index=0,
        key="izbran_izvestaj_sidebar",
        label_visibility="collapsed"
    )
    st.sidebar.markdown(
        """
        <div class="sigma-sidebar-footer">
            <b>PREMIUM модул</b><br>
            Извештаи, KPI, Cash Flow и Valuation на едно место.
        </div>
        """,
        unsafe_allow_html=True
    )
else:
    izbran_izvestaj = "🏠 Заклучен лист"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

if os.path.exists(os.path.join(BASE_DIR, "data", "pravila bilansi.xlsx")):
    pravila_file = os.path.join(BASE_DIR, "data", "pravila bilansi.xlsx")
else:
    pravila_file = os.path.join(BASE_DIR, "pravila bilansi.xlsx")



def read_excel_fill_merged(file, engine="openpyxl", sheet_name=0):
    """
    Чита Excel и ги пополнува merged cells со вредноста од горната-лева ќелија.
    Ова помага кај заклучни листи каде конто/назив или заглавија се споени.
    """
    if engine != "openpyxl":
        file.seek(0)
        return pd.read_excel(file, engine=engine, header=None).dropna(how="all")

    try:
        from openpyxl import load_workbook
        file.seek(0)
        wb = load_workbook(file, data_only=True)
        ws = wb[wb.sheetnames[sheet_name] if isinstance(sheet_name, int) else sheet_name]

        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = ws.cell(min_row, min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row, col).value = value

        data = list(ws.values)
        return pd.DataFrame(data).dropna(how="all")
    except Exception:
        file.seek(0)
        return pd.read_excel(file, engine=engine, header=None).dropna(how="all")


def validate_zaklucen_list(df):
    """
    Контрола дали заклучниот лист е затворен:
    - претходна должи = претходна побарува
    - тековна должи = тековна побарува
    - вкупно должи = вкупно побарува
    - крајно должи = крајно побарува
    """
    checks = []

    pairs = [
        ("Prethodna godina", "prethodna_dolzi", "prethodna_pobaruva"),
        ("Tekovna godina", "tekovna_dolzi", "tekovna_pobaruva"),
        ("Vkupno", "vkupno_dolzi", "vkupno_pobaruva"),
        ("Krajno saldo", "krajno_dolzi", "krajno_pobaruva"),
    ]

    for label, debit_col, credit_col in pairs:
        if debit_col in df.columns and credit_col in df.columns:
            debit = float(df[debit_col].sum())
            credit = float(df[credit_col].sum())
            checks.append({
                "Kontrola": label,
                "Dolzi": round(debit, 2),
                "Pobaruva": round(credit, 2),
                "Razlika": round(debit - credit, 2),
                "Status": "✅ OK" if round(debit - credit, 2) == 0 else "❌ Razlika"
            })

    return pd.DataFrame(checks)


def format_aop(x):
    try:
        if pd.isna(x):
            return ""
        s = str(x).strip()
        if s.endswith(".0"):
            s = s[:-2]
        if s.isdigit():
            return s.zfill(3)
        return s
    except Exception:
        return str(x).strip()


def normalize_key(x):
    if pd.isna(x):
        return ""
    return str(x).strip().replace(" ", "").replace(".0", "").upper()

def pdf_number(x):
    """Gi cita i US (92,250.00) i EU/MK (400.000,00) formati."""
    try:
        s = str(x).strip().replace(" ", "")
        if s == "" or s.lower() == "nan":
            return 0.0

        # Ako ima i tocka i zapirka, posledniot separator e decimalen separator.
        if "." in s and "," in s:
            if s.rfind(",") > s.rfind("."):
                # 400.000,00 -> 400000.00
                s = s.replace(".", "").replace(",", ".")
            else:
                # 92,250.00 -> 92250.00
                s = s.replace(",", "")
        elif "," in s:
            # 1234,56 -> 1234.56 ; 1,234 -> 1234 ako e thousands
            if re.match(r"^-?\d{1,3}(,\d{3})+$", s):
                s = s.replace(",", "")
            else:
                s = s.replace(",", ".")
        return float(s)
    except:
        return 0.0


def split_signed_saldo(value):
    """Ako saldoto e vo edna kolona: plus = dolzi, minus = pobaruva."""
    value = pdf_number(value)
    if value >= 0:
        return value, 0.0
    return 0.0, abs(value)


def extract_pdf_konto_naziv(left_part):
    """Poddrzuva konta kako 002, 10009, 310 01, 742 GP1, 741 U1."""
    left_part = str(left_part).strip()
    match = re.match(r"^([0-9]{3,6}(?:\s+[A-Za-zА-Ша-ш0-9]{1,4})?)\s*(.*)$", left_part)
    if not match:
        return None, None
    konto = match.group(1).replace(" ", "").strip()
    naziv = match.group(2).strip()
    return konto, naziv


def _pdf_number_pattern():
    # US: 92,250.00 / EU: 400.000,00 / plain: 1000.00 ili 1000,00
    return r"-?(?:\d{1,3}(?:[\.,]\d{3})+[\.,]\d{2}|\d+[\.,]\d{2})"


def read_zaklucen_pdf_words_format(pdf):
    """
    Format od drug softver: koloni so X-pozicii:
    Pocetna D/P, Tekoven promet D/P, Vkupen promet D/P, Saldo D/P.
    Ovoj parser gi cita i redovite kade sto praznite nuli ne se ispecateni.
    """
    rows = []
    num_re = re.compile(_pdf_number_pattern())

    # centri na 8-te brojceni koloni vo ovoj landscape PDF format
    column_centers = [274, 342, 411, 480, 548, 617, 685, 754]
    col_names = [
        "prethodna_dolzi", "prethodna_pobaruva",
        "tekovna_dolzi", "tekovna_pobaruva",
        "vkupno_dolzi", "vkupno_pobaruva",
        "krajno_dolzi", "krajno_pobaruva"
    ]

    for page in pdf.pages:
        words = page.extract_words(x_tolerance=2, y_tolerance=3)
        lines = {}
        for w in words:
            top = round(w.get("top", 0))
            lines.setdefault(top, []).append(w)

        for top in sorted(lines):
            ws = sorted(lines[top], key=lambda w: w["x0"])
            if not ws:
                continue

            first = ws[0]["text"].strip()
            # gi zemame samo sintetski/analiticki konta, ne grupni zbirni redovi 0,1,2... i Vкупно
            if not re.fullmatch(r"^\d{6}$", first):
                continue

            konto = first
            text_words = []
            values = {c: 0.0 for c in col_names}

            for w in ws[1:]:
                txt = w["text"].strip()
                if num_re.fullmatch(txt):
                    center = (w["x0"] + w["x1"]) / 2
                    idx = min(range(len(column_centers)), key=lambda i: abs(center - column_centers[i]))
                    values[col_names[idx]] += pdf_number(txt)
                elif w["x0"] < 240:
                    text_words.append(txt)

            # mora da ima barem edna brojka vo redot
            if sum(abs(v) for v in values.values()) == 0:
                continue

            rows.append({
                "konto": konto,
                "naziv": " ".join(text_words).strip(),
                **values
            })

    return pd.DataFrame(rows)



# =====================================================
# PDF PARSERS: IVA SOFT и CFMA / Info Biro
# Додадени како посебни parser-и за да не се расипе старата логика.
# =====================================================

def read_pdf_ivasoft_format(pdf):
    """
    IVA SOFT / BBILANS формат.

    Важно:
    - Кај IVA SOFT ги земаме само аналитичките 6-цифрени конта.
    - Синтетичките 3-цифрени редови ги прескокнуваме.
    - Овој PDF визуелно има бројки малку под конто редот, затоа редот се
      составува по anchor-конто и X-позиции, а не само по ист top/y ред.
    - САЛДО е една signed колона: + во krajno_dolzi, - во krajno_pobaruva.
    """
    rows = []
    num_re = re.compile(_pdf_number_pattern())

    col_centers = [196, 295, 394, 493, 592, 691, 790]
    col_names = [
        "prethodna_dolzi", "prethodna_pobaruva",
        "tekovna_dolzi", "tekovna_pobaruva",
        "vkupno_dolzi", "vkupno_pobaruva",
        "saldo"
    ]

    for page in pdf.pages:
        words = page.extract_words(
            x_tolerance=2,
            y_tolerance=3,
            keep_blank_chars=False
        )

        if not words:
            continue

        words = sorted(words, key=lambda w: (w.get("top", 0), w.get("x0", 0)))

        # Anchor редови = зборови во првата колона кои се конта.
        # Ги земаме само 6-цифрените конта, а 3-цифрените синтетики ги игнорираме.
        account_anchors = []
        for w in words:
            txt = str(w.get("text", "")).strip()
            x0 = float(w.get("x0", 0))
            if x0 < 70 and re.fullmatch(r"\d{3,6}", txt):
                account_anchors.append(w)

        for i, acc in enumerate(account_anchors):
            konto = str(acc.get("text", "")).strip()

            # IVA SOFT: аналитиката е 6 цифри, синтетиката е 3 цифри.
            if not re.fullmatch(r"\d{6}", konto):
                continue

            top_start = float(acc.get("top", 0)) - 2

            # До следното конто-редче. Така ги фаќаме и бројките кои се
            # 3-5 пиксели под конто редот, но не влегуваме во следната сметка.
            if i + 1 < len(account_anchors):
                top_end = float(account_anchors[i + 1].get("top", 0)) - 1
            else:
                top_end = float(acc.get("bottom", acc.get("top", 0))) + 18

            row_words = [
                w for w in words
                if top_start <= float(w.get("top", 0)) < top_end
            ]

            values = {c: 0.0 for c in col_names}
            text_words = []

            for w in row_words:
                txt = str(w.get("text", "")).strip()

                if txt == konto:
                    continue

                center = (float(w.get("x0", 0)) + float(w.get("x1", 0))) / 2

                # Бројки ги земаме само од бројчените колони.
                # Ова спречува опис како "TIP 40.10" да се прочита како износ.
                if center > 140 and num_re.fullmatch(txt):
                    idx = min(
                        range(len(col_centers)),
                        key=lambda j: abs(center - col_centers[j])
                    )
                    values[col_names[idx]] += pdf_number(txt)

                # Описот е меѓу конто колоната и првата бројчена колона.
                elif 45 <= float(w.get("x0", 0)) < 170:
                    text_words.append(txt)

            if sum(abs(v) for v in values.values()) == 0:
                continue

            saldo = values.pop("saldo")
            krajno_dolzi, krajno_pobaruva = split_signed_saldo(saldo)

            rows.append({
                "konto": konto,
                "naziv": " ".join(text_words).strip(),
                "prethodna_dolzi": values["prethodna_dolzi"],
                "prethodna_pobaruva": values["prethodna_pobaruva"],
                "tekovna_dolzi": values["tekovna_dolzi"],
                "tekovna_pobaruva": values["tekovna_pobaruva"],
                "vkupno_dolzi": values["vkupno_dolzi"],
                "vkupno_pobaruva": values["vkupno_pobaruva"],
                "krajno_dolzi": krajno_dolzi,
                "krajno_pobaruva": krajno_pobaruva,
            })

    return normalize_dataframe_numbers(pd.DataFrame(rows))


def read_pdf_cfma_infobiro_format(pdf):
    """
    CFMA / Info Biro BRUTO BILANS формат.

    Ги чита 3-цифрените синтетички редови бидејќи тие се најстабилни
    за AOP пресметки и не создаваат дуплирање со аналитиките.
    """
    rows = []
    num_re = re.compile(_pdf_number_pattern())

    col_centers = [276, 344, 414, 482, 552, 620, 690, 758]
    col_names = [
        "prethodna_dolzi", "prethodna_pobaruva",
        "tekovna_dolzi", "tekovna_pobaruva",
        "vkupno_dolzi", "vkupno_pobaruva",
        "krajno_dolzi", "krajno_pobaruva",
    ]

    for page in pdf.pages:
        words = page.extract_words(
            x_tolerance=2,
            y_tolerance=3,
            keep_blank_chars=False
        )

        lines = {}
        for w in words:
            top = round(w.get("top", 0), 1)
            lines.setdefault(top, []).append(w)

        for top in sorted(lines):
            ws = sorted(lines[top], key=lambda w: w["x0"])
            if not ws:
                continue

            first = ws[0]["text"].strip()

            # CFMA: ги земаме синтетичките 3-цифрени редови
            if not re.fullmatch(r"\d{3}", first):
                continue

            # Кај CFMA конто колоната е лево
            if ws[0].get("x0", 0) > 60:
                continue

            values = {c: 0.0 for c in col_names}
            text_words = []

            for w in ws[1:]:
                txt = w["text"].strip()
                if num_re.fullmatch(txt):
                    center = (w["x0"] + w["x1"]) / 2
                    idx = min(
                        range(len(col_centers)),
                        key=lambda i: abs(center - col_centers[i])
                    )
                    values[col_names[idx]] += pdf_number(txt)
                elif w.get("x0", 0) < 240:
                    text_words.append(txt)

            if sum(abs(v) for v in values.values()) == 0:
                continue

            rows.append({
                "konto": first,
                "naziv": " ".join(text_words).strip(),
                **values
            })

    return normalize_dataframe_numbers(pd.DataFrame(rows))

def _pdf_balance_diff(df):
    """
    Pomoshna kontrola za izbor na najdobar PDF parser.
    Kolku e pomal zbirniot debet/kredit diff, tolku e parserot posiguren.
    """
    if df is None or df.empty:
        return float("inf")

    required_cols = [
        "prethodna_dolzi", "prethodna_pobaruva",
        "tekovna_dolzi", "tekovna_pobaruva",
        "vkupno_dolzi", "vkupno_pobaruva",
        "krajno_dolzi", "krajno_pobaruva"
    ]

    if not all(c in df.columns for c in required_cols):
        return float("inf")

    diff = 0.0
    pairs = [
        ("prethodna_dolzi", "prethodna_pobaruva"),
        ("tekovna_dolzi", "tekovna_pobaruva"),
        ("vkupno_dolzi", "vkupno_pobaruva"),
        ("krajno_dolzi", "krajno_pobaruva"),
    ]

    for d_col, p_col in pairs:
        d = float(df[d_col].sum())
        p = float(df[p_col].sum())
        diff += abs(round(d - p, 2))

    return diff


def _pdf_parser_score(df, priority):
    """
    Score za avtomatski izbor:
    1) prvo balansirana struktura,
    2) potoa povekje procitani redovi,
    3) potoa prioritet na parser.
    """
    if df is None or df.empty or len(df) <= 5:
        return (-1, float("inf"), 0, -priority)

    balance_diff = _pdf_balance_diff(df)
    is_balanced = 1 if balance_diff <= 1.0 else 0

    return (is_balanced, -balance_diff, len(df), -priority)


def read_zaklucen_pdf_standard_lines(pdf):
    """
    Standarden parser po tekst-linija.
    Go koristime kako kandidat, a ne kako posledna opcija,
    za da ne se pogreshat stari Zonel bruto bilansi kako IVA SOFT.
    """
    rows = []
    number_pattern = _pdf_number_pattern()

    for page in pdf.pages:
        text = page.extract_text()

        if not text:
            continue

        for line in text.split("\n"):
            line = line.strip()
            nums = re.findall(number_pattern, line)

            # Standarden PDF: pocetna D/P + tekovna D/P + vkupno D/P + krajno D/P = 8 brojki
            # Obicen PDF: pocetna saldo + tekovna D/P + vkupno D/P + krajno saldo = 6 brojki
            if len(nums) >= 8 or len(nums) == 6:

                first_num = re.search(number_pattern, line)
                if not first_num:
                    continue

                left_part = line[:first_num.start()].strip()
                konto, naziv = extract_pdf_konto_naziv(left_part)

                if not konto:
                    continue

                if len(nums) >= 8:
                    prethodna_dolzi = pdf_number(nums[0])
                    prethodna_pobaruva = pdf_number(nums[1])
                    tekovna_dolzi = pdf_number(nums[2])
                    tekovna_pobaruva = pdf_number(nums[3])
                    vkupno_dolzi = pdf_number(nums[4])
                    vkupno_pobaruva = pdf_number(nums[5])
                    krajno_dolzi = pdf_number(nums[6])
                    krajno_pobaruva = pdf_number(nums[7])
                else:
                    # Format so pocetno i krajno saldo vo edna kolona
                    prethodna_dolzi, prethodna_pobaruva = split_signed_saldo(nums[0])
                    tekovna_dolzi = pdf_number(nums[1])
                    tekovna_pobaruva = pdf_number(nums[2])
                    vkupno_dolzi = pdf_number(nums[3])
                    vkupno_pobaruva = pdf_number(nums[4])
                    krajno_dolzi, krajno_pobaruva = split_signed_saldo(nums[5])

                rows.append({
                    "konto": konto,
                    "naziv": naziv,
                    "prethodna_dolzi": prethodna_dolzi,
                    "prethodna_pobaruva": prethodna_pobaruva,
                    "tekovna_dolzi": tekovna_dolzi,
                    "tekovna_pobaruva": tekovna_pobaruva,
                    "vkupno_dolzi": vkupno_dolzi,
                    "vkupno_pobaruva": vkupno_pobaruva,
                    "krajno_dolzi": krajno_dolzi,
                    "krajno_pobaruva": krajno_pobaruva
                })

    return normalize_dataframe_numbers(pd.DataFrame(rows))


def _read_zaklucen_pdf_legacy_unused(file):
    st.success("PDF uspešno učitan. Počnuvam so ekstrakcija na podatoci...")

    with pdfplumber.open(file) as pdf:

        text_all = "\n".join([p.extract_text() or "" for p in pdf.pages])
        text_upper = text_all.upper()

        candidates = []

        # 1) STANDARD parser prvo kako kandidat.
        # Ova e bitno za starite Zonel bruto bilansi koi vo naslov mozat da imaat "BRUTO BILANS",
        # no strukturata im e klasicna so 8 debet/kredit koloni.
        df_standard = read_zaklucen_pdf_standard_lines(pdf)
        if not df_standard.empty and len(df_standard) > 5:
            candidates.append({
                "name": "STANDARD / ZONEL struktura",
                "df": df_standard,
                "priority": 1
            })

        # 2) Parser po X-pozicii za format kade praznite nuli ne se pecatat.
        # Go probuvame spored struktura/tekst, no izborot pak go pravi score-ot.
        if "ZAKLU" in text_upper and "SALDO" in text_upper and "TEKOV" in text_upper:
            df_words = read_zaklucen_pdf_words_format(pdf)
            if not df_words.empty and len(df_words) > 5:
                candidates.append({
                    "name": "Zaklucen list so prazni koloni bez nuli / X-pozicii",
                    "df": df_words,
                    "priority": 2
                })

        # 3) CFMA / Info Biro kandidat.
        if "BRUTO BILANS" in text_upper and ("KONTO" in text_upper or "PURE" in text_upper):
            df_cfma = read_pdf_cfma_infobiro_format(pdf)
            if not df_cfma.empty and len(df_cfma) > 5:
                candidates.append({
                    "name": "CFMA / Info Biro BRUTO BILANS",
                    "df": df_cfma,
                    "priority": 3
                })

        # 4) IVA SOFT kandidat.
        # Vazno: ne go vrakjame avtomatski samo zatoa shto vo tekstot pisuva "БРУТОБИЛАНС".
        # Prvo go sporeduvame so standardniot parser po realna struktura i kontrola.
        if "БРУТОБИЛАНС" in text_upper or "БРУТО БИЛАНС" in text_upper:
            df_iva = read_pdf_ivasoft_format(pdf)
            if not df_iva.empty and len(df_iva) > 5:
                candidates.append({
                    "name": "IVA SOFT / BBILANS",
                    "df": df_iva,
                    "priority": 4
                })

        if candidates:
            best = sorted(
                candidates,
                key=lambda c: _pdf_parser_score(c["df"], c["priority"]),
                reverse=True
            )[0]

            best_df = normalize_dataframe_numbers(best["df"])
            balance_diff = _pdf_balance_diff(best_df)

            st.info(
                f"Detektiran PDF format: {best['name']} "
                f"(redovi: {len(best_df)}, kontrolna razlika: {balance_diff:,.2f})."
            )

            return best_df

    st.error("PDF e procitan, no ne se pronajdeni redovi od bruto bilansot.")
    st.stop()


def read_zaklucen(file):
    engine = "xlrd" if file.name.lower().endswith(".xls") else "openpyxl"
    raw = read_excel_fill_merged(file, engine=engine, sheet_name=0)

    start_row = None
    for i in range(len(raw)):
        val = str(raw.iloc[i, 0]).strip().lower()
        if val in ["konto", "конто"] or val.isdigit():
            start_row = i
            break

    if start_row is None:
        st.error("Ne moze da se najde pocetok na zaklucniot list.")
        st.stop()

    first_val = str(raw.iloc[start_row, 0]).strip().lower()
    if first_val in ["konto", "конто"]:
        start_row += 1

    df = raw.iloc[start_row:].copy()

    if df.shape[1] >= 10:
        df = df.iloc[:, :10]
        df.columns = [
            "konto", "naziv",
            "prethodna_dolzi", "prethodna_pobaruva",
            "tekovna_dolzi", "tekovna_pobaruva",
            "vkupno_dolzi", "vkupno_pobaruva",
            "krajno_dolzi", "krajno_pobaruva"
        ]
    else:
        df = df.iloc[:, :9]
        df.columns = [
            "konto", "naziv", "saldo",
            "prethodna_dolzi", "prethodna_pobaruva",
            "tekovna_dolzi", "tekovna_pobaruva",
            "vkupno_dolzi", "vkupno_pobaruva"
        ]
        df["saldo"] = df["saldo"].apply(clean_number)
        df["krajno_dolzi"] = df["saldo"].apply(lambda x: x if x > 0 else 0)
        df["krajno_pobaruva"] = df["saldo"].apply(lambda x: abs(x) if x < 0 else 0)

    df["konto"] = df["konto"].astype(str).str.replace(".0", "", regex=False).str.strip()
    df["naziv"] = df["naziv"].astype(str)

    for col in df.columns:
        if col not in ["konto", "naziv"]:
            df[col] = df[col].apply(clean_number)

    df = normalize_dataframe_numbers(df)
    return df[(df["konto"] != "") & (df["konto"] != "nan")]


def find_header_row(file, sheet_name):
    temp = pd.read_excel(file, sheet_name=sheet_name, header=None)
    for i in range(len(temp)):
        row = [str(x).strip().lower() for x in temp.iloc[i].tolist()]
        if "aop" in row:
            return i
    raise ValueError(f"Ne ja najdov kolonata AOP vo sheet: {sheet_name}")


@st.cache_data(show_spinner=False, max_entries=32)
def read_rules(file, sheet_name):
    # Правилата се статични во рамки на deploy; нема потреба повторно да се
    # чита Excel sheet при секој Streamlit rerun.
    header_row = find_header_row(file, sheet_name)
    rules = pd.read_excel(file, sheet_name=sheet_name, header=header_row)
    rules.columns = [str(c).strip() for c in rules.columns]

    rename = {}
    for col in rules.columns:
        c = str(col).strip().lower()
        if c in ["aop", "ознака на аоп"]:
            rename[col] = "AOP"
        elif c in ["pozicija", "позиција"]:
            rename[col] = "Pozicija"
        elif c == "naziv":
            rename[col] = "Naziv"
        elif c == "opis":
            rename[col] = "Opis"
        elif c == "tip":
            rename[col] = "Tip"
        elif c == "izvor":
            rename[col] = "Izvor"
        elif c == "key":
            rename[col] = "Key"
        elif c == "kategorija":
            rename[col] = "Kategorija"
        elif c == "konto":
            rename[col] = "konto"
        elif c == "konto.1":
            rename[col] = "konto_prethodna"
        elif c == "kolona":
            rename[col] = "kolona"
        elif c == "kolona.1":
            rename[col] = "kolona_prethodna"
        elif c == "operacija":
            rename[col] = "Operacija"
        elif c == "operacija.1":
            rename[col] = "Operacija_prethodna"
        elif c == "formula":
            rename[col] = "Formula"
        elif c == "logika":
            rename[col] = "Logika"

    rules = rules.rename(columns=rename)
    if "AOP" not in rules.columns:
        raise ValueError(f"Vo {sheet_name} ne postoi kolona AOP.")

    rules = rules[rules["AOP"].notna()].copy()
    rules["AOP"] = rules["AOP"].apply(format_aop)
    return rules


def konto_list(konto_text):
    if pd.isna(konto_text):
        return []
    text = str(konto_text).replace("(", "").replace(")", "").replace(" ", "").replace(";", ",")
    result = []
    for part in text.split(","):
        if not part:
            continue
        if "-" in part:
            try:
                a, b = part.split("-")
                width = max(len(a), len(b))
                for n in range(int(a), int(b) + 1):
                    result.append(str(n).zfill(width))
            except Exception:
                st.error(e)
                pass
        else:
            result.append(part)
    return result


def sum_rule(df, konto_text, kolona, operacija):
    if pd.isna(konto_text) or pd.isna(kolona):
        return 0.0
    kolona = str(kolona).strip()
    if kolona not in df.columns:
        return 0.0
    konta = konto_list(konto_text)
    if not konta:
        return 0.0
    mask = df["konto"].astype(str).str.startswith(tuple(konta), na=False)
    value = df.loc[mask, kolona].sum()
    if str(operacija).strip() == "-":
        return -value
    return value


def safe_eval_expression(expr):
    allowed_nodes = (
        ast.Expression, ast.BinOp, ast.UnaryOp,
        ast.Add, ast.Sub, ast.Mult, ast.Div,
        ast.USub, ast.UAdd,
        ast.Constant, ast.Load, ast.Expr
    )
    tree = ast.parse(expr, mode="eval")
    for node in ast.walk(tree):
        if not isinstance(node, allowed_nodes):
            raise ValueError(f"Invalid formula node: {type(node).__name__}")
    return eval(compile(tree, "<string>", "eval"))


def replace_formula_tokens(formula, values, external_values=None):
    """
    Поддржува формули со:
    - обични AOP: 063+111
    - BU референци: BU246, BU247
    - BS референци: BS063 или BS063_T / BS063_P / BS063_R
    """
    external_values = external_values or {}
    all_values = {}
    all_values.update({str(k).upper(): v for k, v in external_values.items()})
    all_values.update({str(k).upper().zfill(3) if str(k).isdigit() else str(k).upper(): v for k, v in values.items()})

    expression = str(formula).strip().replace("=", "").replace(" ", "").upper()

    # Прво замени подолги клучеви како BS063_T, потоа BU246, потоа 063
    for key in sorted(all_values.keys(), key=len, reverse=True):
        expression = re.sub(rf"\b{re.escape(key)}\b", str(all_values.get(key, 0.0)), expression)

    # ВАЖНО:
    # Не ги заменуваме повторно сите 3-цифрени броеви после првата замена.
    # Причина: кога формулата 225+226+227+228 ќе се претвори во 0.0+0.0+394.0+0.0,
    # regex-от може погрешно да го препознае бројот 394 од 394.0 како AOP 394 и да го замени со 0.
    # Затоа тука враќаме expression директно.
    return expression


def apply_formulas_and_logic(rules, values, external_values=None):
    values = values.copy()
    for _ in range(30):
        for _, row in rules.iterrows():
            aop = format_aop(row["AOP"])
            formula = row.get("Formula", "")
            if pd.isna(formula) or str(formula).strip() == "":
                continue


            try:
                expression = replace_formula_tokens(formula, values, external_values)
                values[aop] = float(safe_eval_expression(expression))
            except Exception:
                values[aop] = 0.0

    for _, row in rules.iterrows():
        aop = format_aop(row["AOP"])
        logic = str(row.get("Logika", "")).strip().lower()
        value = values.get(aop, 0.0)
        if logic == "positive":
            values[aop] = value if value > 0 else 0.0
        elif logic == "negative":
            values[aop] = abs(value) if value < 0 else 0.0
    return values


def calculate_bilans_uspeh(df, rules_file):
    rules = read_rules(rules_file, "Pravila Bilans Uspeh")
    values, positions, order = {}, {}, []
    for _, row in rules.iterrows():
        aop = format_aop(row["AOP"])
        if aop not in order:
            order.append(aop)
        positions[aop] = row.get("Pozicija", "")
        val = sum_rule(df, row.get("konto", None), row.get("kolona", None), row.get("Operacija", "+"))
        values[aop] = values.get(aop, 0.0) + val
    values = apply_formulas_and_logic(rules, values)
    return pd.DataFrame([{"AOP": aop, "Pozicija": positions.get(aop, ""), "Iznos": values.get(aop, 0.0)} for aop in order])


def calculate_seopfatna_dobivka(bu, rules_file):

    rules = pd.read_excel(
        rules_file,
        sheet_name="Pravila Seopfatna Dobivka",
        header=1
    )

    rules.columns = [str(c).strip() for c in rules.columns]
    rules = rules[rules["Red"].notna()].copy()

    data_map = {}

    for _, row in bu.iterrows():
        aop = str(row["AOP"]).strip().replace(".0", "").zfill(3)
        data_map[f"BU{aop}"] = float(row["Iznos"])

    results = {}

    rows = []

    for _, row in rules.iterrows():

        red = str(row["Red"]).strip()
        naziv = str(row["Naziv"]).strip()
        formula = str(row["Formula"]).strip().replace(" ", "")
        tip = str(row.get("Tip", "iznos")).strip()
        stil = str(row.get("Stil", "normal")).strip()

        expression = formula

        all_values = {}
        all_values.update(data_map)
        all_values.update(results)

        for key in sorted(all_values.keys(), key=len, reverse=True):
            expression = expression.replace(key, str(all_values[key]))

        try:
            value = float(safe_eval_expression(expression))
        except:
            value = 0.0

        results[red] = value

        rows.append({
            "Red": red,
            "Naziv": naziv,
            "Tekovna godina": value,
            "Tip": tip,
            "Stil": stil
        })

    return pd.DataFrame(rows)



# =====================================================
# PROPORCIONALNA RASPREDELBA NA ZBIRNA ISPRAVKA 019
# =====================================================

def norm_konto_019(x):
    return (
        str(x)
        .replace("-", "")
        .replace("/", "")
        .replace(".", "")
        .replace(" ", "")
        .strip()
    )


def suma_konto_019_exact(df, konto, kolona):
    """
    Go zema samo sintetickoto konto:
    019, 0190, 019000, 019-000, 019/000...
    No ne gi zema analitikite 0191, 0192, 0193...
    """
    if "konto" not in df.columns or kolona not in df.columns:
        return 0.0

    k = norm_konto_019(konto)
    df_konta = df["konto"].astype(str).apply(norm_konto_019)

    mask = df_konta.str.fullmatch(k + r"0*")
    return float(df.loc[mask, kolona].sum())


def ima_analitika_019(df):
    """
    Vraka True ako postoi 0191, 0192, 0193...
    Tuka 0190 i 019000 NE se smetaat za analitika, tuku za zbirno 019.
    """
    if "konto" not in df.columns:
        return False

    df_konta = df["konto"].astype(str).apply(norm_konto_019)
    return df_konta.str.match(r"019[1-9]").any()


def ima_zbirno_019(df):
    if "konto" not in df.columns:
        return False

    df_konta = df["konto"].astype(str).apply(norm_konto_019)
    return df_konta.str.fullmatch(r"0190*").any()


def raspredeli_019_proporcionalno(df, values, kolona_019, kolona_bruto):
    """
    Ako nema analitika 0191/0192/0193..., a ima samo zbirno 019,
    togash 019 se odzema proporcionalno od site materijalni sredstva
    koi imaat bruto vrednost > 0.

    Vazno:
    - AOP 015 ne sme da ja nosi celata amortizacija 019.
    - Zatoa vo ovoj slucaj AOP 015 se postavuva na 0.
    """
    if not ima_zbirno_019(df):
        return values

    if ima_analitika_019(df):
        return values

    iznos_019 = abs(suma_konto_019_exact(df, "019", kolona_019))

    if iznos_019 == 0:
        return values

    # AOP -> bruto konto
    bruto_map = {
        "012": "011",
        "013": "012",
        "014": "013",
        "016": "014",
        "019": "015",
    }

    bruto_values = {}
    vkupno_bruto = 0.0

    for aop, konto in bruto_map.items():
        bruto = abs(suma_konto_019_exact(df, konto, kolona_bruto))

        if bruto > 0:
            bruto_values[aop] = bruto
            vkupno_bruto += bruto

    if vkupno_bruto == 0:
        return values

    # 019 ne smee celosno da ostane vo AOP 015
    values["015"] = 0.0

    # Odzemi proporcionalen del od 019 od site sredstva so bruto > 0
    for aop, bruto in bruto_values.items():
        del_od_019 = iznos_019 * (bruto / vkupno_bruto)
        values[aop] = values.get(aop, 0.0) - del_od_019

    return values

def calculate_bilans_sostojba(df, rules_file, bu=None):
    rules = read_rules(rules_file, "Pravila bilans Sostojba")
    current_values, previous_values, positions, order = {}, {}, {}, []

    bu_current_map = {}
    if bu is not None:
        for _, bu_row in bu.iterrows():
            bu_aop = str(bu_row["AOP"]).strip().replace(".0", "").zfill(3)
            bu_current_map[f"BU{bu_aop}"] = float(bu_row["Iznos"])

    for _, row in rules.iterrows():
        aop = format_aop(row["AOP"])
        if aop not in order:
            order.append(aop)
        positions[aop] = row.get("Pozicija", "")

                
        current_val = sum_rule(
            df,
            row.get("konto", None),
            row.get("kolona", None),
            row.get("Operacija", "+")
        )

            

        previous_val = sum_rule(
            df,
            row.get("konto_prethodna", row.get("konto", None)),
            row.get("kolona_prethodna", None),
            row.get("Operacija_prethodna", row.get("Operacija", "+"))
        )

        current_values[aop] = current_values.get(aop, 0.0) + current_val
        previous_values[aop] = previous_values.get(aop, 0.0) + previous_val

    # Сметководствено правило:
    # Нето добивка/загуба од БУ НЕ се пополнува како разлика за затворање,
    # туку директно се пренесува од БУ во БС:
    # BU255 -> BS077, BU256 -> BS078.
    # Потоа формулите во БС повторно ги пресметуваат збирните AOP позиции.
    if bu is not None:
        bu255 = bu_current_map.get("BU255", 0.0)
        bu256 = bu_current_map.get("BU256", 0.0)
        current_values["077"] = bu255
        current_values["078"] = bu256

    current_values = raspredeli_019_proporcionalno(
        df,
        current_values,
        "krajno_pobaruva",
        "krajno_dolzi"
    )

    current_values = apply_formulas_and_logic(rules, current_values, external_values=bu_current_map)

    # Сигурност: ако правилата имаат формула на 077/078, повторно форсираме директен пренос
    # и уште еднаш ги освежуваме збирните формули.
    if bu is not None:
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)
        current_values = apply_formulas_and_logic(rules, current_values, external_values=bu_current_map)
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)

    previous_values = raspredeli_019_proporcionalno(
        df,
        previous_values,
        "prethodna_pobaruva",
        "prethodna_dolzi"
    )

    previous_values = apply_formulas_and_logic(rules, previous_values)

    return pd.DataFrame([
        {
            "AOP": aop,
            "Pozicija": positions.get(aop, ""),
            "Tekovna godina": current_values.get(aop, 0.0),
            "Prethodna godina": previous_values.get(aop, 0.0)
        }
        for aop in order
    ])



def auto_fix_bs_display_with_bu(bs, bu):
    """
    Безбедна корекција само на готовиот BS dataframe.
    Не го прекинува процесот и не ги менува правилата.
    Ако Активата и Пасивата не се еднакви, ја додава разликата во најсоодветен ред
    за добивка/загуба во капиталот и ја усогласува AOP 111.
    """
    note = None
    try:
        bs = bs.copy()
        aktiva = float(bs.loc[bs["AOP"].astype(str).str.zfill(3) == "063", "Tekovna godina"].sum())
        pasiva = float(bs.loc[bs["AOP"].astype(str).str.zfill(3) == "111", "Tekovna godina"].sum())
        diff = round(aktiva - pasiva, 2)
        if diff == 0:
            return bs, note

        bu_profit = float(bu.loc[bu["AOP"].astype(str).str.zfill(3) == "246", "Iznos"].sum()) if bu is not None else 0.0
        bu_loss = float(bu.loc[bu["AOP"].astype(str).str.zfill(3) == "247", "Iznos"].sum()) if bu is not None else 0.0
        net_bu = round(bu_profit - bu_loss, 2)

        # Ако разликата е приближно еднаква со добивка/загуба од BU, ја внесуваме во капитал.
        # Ако не е еднаква, сепак ја прикажуваме како техничка BS корекција за да не падне извештајот.
        candidates_by_text = []
        for idx, row in bs.iterrows():
            text = str(row.get("Pozicija", "")).lower()
            aop = str(row.get("AOP", "")).strip().replace(".0", "").zfill(3)
            if any(w in text for w in ["добив", "dobiv", "загуб", "zagub", "резултат", "rezultat"]):
                candidates_by_text.append((idx, aop))

        candidate_idx = None
        candidate_aop = None
        for idx, aop in candidates_by_text:
            if aop not in ["063", "111"]:
                candidate_idx, candidate_aop = idx, aop
                break

        if candidate_idx is None:
            for wanted_aop in ["107", "108", "109", "110", "095"]:
                matches = bs.index[bs["AOP"].astype(str).str.replace(".0", "", regex=False).str.zfill(3) == wanted_aop].tolist()
                if matches:
                    candidate_idx = matches[0]
                    candidate_aop = wanted_aop
                    break

        if candidate_idx is None:
            return bs, {
                "status": "warning",
                "message": f"BS ne e zatvoren. Razlika: {diff:,.0f}. Ne najdov red za avtomatska korekcija na dobivka/zaguba."
            }

        bs.loc[candidate_idx, "Tekovna godina"] = float(bs.loc[candidate_idx, "Tekovna godina"]) + diff

        idx_111 = bs.index[bs["AOP"].astype(str).str.replace(".0", "", regex=False).str.zfill(3) == "111"].tolist()
        if idx_111:
            bs.loc[idx_111[0], "Tekovna godina"] = float(bs.loc[idx_111[0], "Tekovna godina"]) + diff

        note = {
            "status": "fixed",
            "aop": candidate_aop,
            "adjustment": diff,
            "old_diff": diff,
            "bu_profit": bu_profit,
            "bu_loss": bu_loss,
            "net_bu": net_bu,
        }
        return bs, note
    except Exception as e:
        return bs, {
            "status": "error",
            "message": f"BS korekcijata ne se primeni: {e}"
        }


def build_data_map(bu, bs):
    data_map = {}
    for _, row in bu.iterrows():
        aop = str(row["AOP"]).strip().replace(".0", "").zfill(3)
        data_map[f"BU{aop}"] = row["Iznos"]
    for _, row in bs.iterrows():
        aop = str(row["AOP"]).strip().replace(".0", "").zfill(3)
        tekovna = row["Tekovna godina"]
        prethodna = row["Prethodna godina"]
        data_map[f"BS{aop}_T"] = tekovna
        data_map[f"BS{aop}_P"] = prethodna
        data_map[f"BS{aop}_R"] = tekovna - prethodna
    return data_map


def calculate_cash_flow(df, rules_file, bu, bs):
    rules = read_rules(rules_file, "Pravila CF")
    data_map = build_data_map(bu, bs)
    values, positions, order, formula_rows = {}, {}, [], []
    for _, row in rules.iterrows():
        aop = normalize_key(row.get("AOP", ""))
        if aop not in order:
            order.append(aop)
        positions[aop] = row.get("Pozicija", "")
        key = normalize_key(row.get("Key", ""))
        operacija = str(row.get("Operacija", "+")).strip()
        formula = str(row.get("Formula", "")).strip()
        value = 0.0
        if key != "" and key.lower() != "nan":
            value = data_map.get(key, 0.0)
        elif formula != "" and formula.lower() != "nan":
            formula_rows.append(row)
        if operacija == "-":
            value = -value
        values[aop] = values.get(aop, 0.0) + value

    for _ in range(20):
        for row in formula_rows:
            aop = normalize_key(row.get("AOP", ""))
            formula = str(row.get("Formula", "")).strip().replace(" ", "").replace("=", "").upper()
            def repl(match):
                ref = normalize_key(match.group(0))
                return str(values.get(ref, 0.0))
            expression = re.sub(r"CF\d+", repl, formula)
            try:
                values[aop] = float(safe_eval_expression(expression))
            except Exception:
                values[aop] = 0.0
    return pd.DataFrame([{"AOP": aop, "Pozicija": positions.get(aop, ""), "Iznos": values.get(aop, 0.0)} for aop in order])


def calculate_kpi(rules_file, bu, bs, broj_vraboteni=0, meseci_rabotenje=0):
    rules = read_rules(rules_file, "Pravila KPI")
    data_map = build_data_map(bu, bs)
    data_map["BROJ_VRABOTENI"] = broj_vraboteni
    data_map["MESECI_RABOTENJE"] = meseci_rabotenje

    results = []

    for _, row in rules.iterrows():
        kpi = str(row.get("AOP", "")).strip()
        naziv = str(row.get("Naziv", "")).strip()
        opis = str(row.get("Opis", "")).strip()
        tip = str(row.get("Tip", "")).strip().lower()
        formula = str(row.get("Formula", "")).strip().replace(" ", "")

        expression = formula

        for key in sorted(data_map.keys(), key=len, reverse=True):
            expression = re.sub(
                rf"\b{re.escape(key)}\b",
                str(data_map[key]),
                expression
            )

        try:
            value = float(eval(expression))
        except Exception:
            value = 0.0

        if tip == "procent":
            prikaz = f"{value * 100:.2f}%"
        elif tip == "iznos":
            prikaz = f"{value:,.0f}"
        else:
            prikaz = f"{value:,.2f}"

        dobro_od = row.get("Dobro_od", None)
        rizik_pod = row.get("Rizik_pod", None)

        status = "ℹ️"

        try:
            dobro_od = float(dobro_od)
            rizik_pod = float(rizik_pod)

            if value >= dobro_od:
                status = "🟢 Добро"
            elif value <= rizik_pod:
                status = "🔴 Ризик"
            else:
                status = "🟡 Средно"

        except:
            pass

        results.append({
            "AOP": kpi,
            "Naziv": naziv,
            "Kategorija": row.get("Kategorija", ""),
            "Opis": opis,
            "Status": status,
            "Vrednost": prikaz,
            "RawValue": value,

        })

    return pd.DataFrame(results)

def calculate_break_even(df, bu, sales_aops, pct_402=70, pct_403=70, months_in_period=12, include_701=True, pct_701=100):
    def sum_konto_prefix(prefix, exclude_accounts=None):
        """
        Break-even сумирање по конто-префикс.

        v8:
        - користи central effective-account matching ако е достапен;
        - затоа не дуплира синтетика + аналитика;
        - за класа 4 ги исклучува 4900/4901 како распоредни конта;
        - 701 може да се третира како варијабилен трошок за трговски фирми.
        """
        exclude_accounts = exclude_accounts or []
        work_df = df.copy()
        konto_series = work_df["konto"].astype(str).str.strip()

        if exclude_accounts:
            work_df = work_df.loc[~konto_series.isin(exclude_accounts)].copy()

        # Во v6/v7 веќе имаме стабилна central effective логика.
        # Ја користиме и тука за да нема дуплирање синтетика/аналитика.
        if "sum_rule_effective" in globals():
            try:
                return float(sum_rule_effective(work_df, str(prefix), "tekovna_dolzi", "+"))
            except Exception:
                pass

        # Fallback ако некоја стара верзија се стартува без effective engine.
        konto_series = work_df["konto"].astype(str).str.strip()
        mask = konto_series.str.startswith(str(prefix), na=False)
        return float(work_df.loc[mask, "tekovna_dolzi"].sum())

    def sum_bu_aops(aops):
        total = 0.0
        for aop in aops:
            aop = str(aop).strip().zfill(3)
            total += float(
                bu.loc[bu["AOP"].astype(str).str.zfill(3) == aop, "Iznos"].sum()
            )
        return total

    sales_revenue = sum_bu_aops(sales_aops)

    total_class_4 = sum_konto_prefix("4", exclude_accounts=["4900", "4901"])

    konto_400 = sum_konto_prefix("400", exclude_accounts=["4900", "4901"])
    konto_402 = sum_konto_prefix("402", exclude_accounts=["4900", "4901"]) * (pct_402 / 100)
    konto_403 = sum_konto_prefix("403", exclude_accounts=["4900", "4901"]) * (pct_403 / 100)

    # 701 = набавна вредност на продадени стоки.
    # Кај трговски фирми ова е најчесто 100% варијабилен трошок.
    konto_701_total = sum_konto_prefix("701") if include_701 else 0.0
    konto_701_variable = konto_701_total * (pct_701 / 100)

    total_costs_for_be = total_class_4 + konto_701_total
    variable_costs = konto_400 + konto_402 + konto_403 + konto_701_variable
    fixed_costs = total_costs_for_be - variable_costs

    variable_ratio = variable_costs / sales_revenue if sales_revenue else 0
    contribution_margin_ratio = 1 - variable_ratio

    break_even_sales = (
        fixed_costs / contribution_margin_ratio
        if contribution_margin_ratio > 0
        else 0
    )

    monthly_break_even = (
        break_even_sales / months_in_period
        if months_in_period > 0
        else 0
    )

    margin_of_safety = sales_revenue - break_even_sales
    margin_of_safety_pct = margin_of_safety / sales_revenue if sales_revenue else 0

    if margin_of_safety_pct >= 0.30:
        status = "🟢 Низок ризик"
        comment = "Компанијата работи со добра сигурносна маргина над break-even точката."
    elif margin_of_safety_pct >= 0.10:
        status = "🟡 Среден ризик"
        comment = "Компанијата е над break-even точката, но потребно е редовно следење на продажбата и трошоците."
    else:
        status = "🔴 Висок ризик"
        comment = "Компанијата работи блиску до break-even точката и е чувствителна на пад на продажбата."

    return pd.DataFrame([
        {"Показател": "Приход од продажба", "Вредност": sales_revenue},
        {"Показател": "Набавна вредност 701", "Вредност": konto_701_total},
        {"Показател": "Варијабилен дел од 701", "Вредност": konto_701_variable},
        {"Показател": "Вкупни трошоци земени во BE", "Вредност": total_costs_for_be},
        {"Показател": "Варијабилни трошоци", "Вредност": variable_costs},
        {"Показател": "Фиксни трошоци", "Вредност": fixed_costs},
        {"Показател": "Варијабилни трошоци %", "Вредност": variable_ratio},
        {"Показател": "Маржа за покривање на фиксни трошоци", "Вредност": contribution_margin_ratio},
        {"Показател": "Break-even продажба", "Вредност": break_even_sales},
        {
        "Показател": "Просечен месечен Break-even",
        "Вредност": monthly_break_even
        },
        {"Показател": "Над Break-even", "Вредност": margin_of_safety},
        {"Показател": "Безбедносна маргина %", "Вредност": margin_of_safety_pct},
        {"Показател": "Статус", "Вредност": status},
        {"Показател": "Коментар", "Вредност": comment},
    ])

def export_excel(sheets):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, data in sheets.items():
            if data is not None:
                data.to_excel(writer, index=False, sheet_name=name[:31])
    return output.getvalue()


def style_formula_rows(row):
    formula_rows = ["201", "207", "213", "223", "224", "234", "235", "246", "247", "250", "251", "001", "002", "009", "010", "020", "021", "031", "036", "037", "045", "052", "053", "059", "063", "065", "071", "081", "082", "085", "095", "111"]
    if str(row["AOP"]).zfill(3) in formula_rows:
        return ["font-weight: bold; font-size: 18px;" for _ in row]
    return ["" for _ in row]


def style_cashflow_rows(row):
    total_rows = ["CF08", "CF12", "CF16", "CF17", "CF20"]
    if row["AOP"] in total_rows:
        if row["AOP"] == "CF17":
            return ["background-color: #2F65D9; color: white; font-weight: 700;" for _ in row]
        return ["background-color: #D9EAD3; font-weight: 700;" for _ in row]
    return ["" for _ in row]


def prepare_pdf_dataframe(df):
    """
    Kompaktna PDF verzija: gi trga redovite kade sto site brojceni koloni se 0.
    Excel exportot ostanuva celosen; ova vazi samo za PDF za pecatenje.
    """
    compact = df.copy()

    numeric_cols = []
    for col in compact.columns:
        if col in ["AOP", "Pozicija", "Naziv", "Red", "Kategorija", "Opis", "Status"]:
            continue
        test_values = pd.to_numeric(compact[col], errors="coerce")
        if test_values.notna().any():
            numeric_cols.append(col)
            compact[col] = test_values.fillna(0)

    if numeric_cols:
        mask = compact[numeric_cols].abs().sum(axis=1) != 0
        compact = compact[mask].copy()

    return compact.reset_index(drop=True)


def format_pdf_value(x):
    try:
        if pd.isna(x):
            return ""
        if isinstance(x, (int, float, np.integer, np.floating)):
            return f"{float(x):,.0f}"
        return str(x)
    except Exception:
        return str(x)


def export_single_pdf(df, title, file_title):
    output = BytesIO()

    # Kompakten A4 landscape PDF: pomali margini + pomal font + skrieni nulti redovi
    pdf_df = prepare_pdf_dataframe(df)

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=10,
        leftMargin=10,
        topMargin=10,
        bottomMargin=10,
    )

    styles = getSampleStyleSheet()
    styles["Title"].fontName = "PDF_Font_Bold"
    styles["Title"].fontSize = 11
    styles["Title"].leading = 12

    # VAŽNO: Paragraph tekstot ne ja prima TEXTCOLOR komandata od TableStyle.
    # Zatoa pravime posebni paragraph stilovi za obicni, subtotal i main-total redovi.
    from reportlab.lib.styles import ParagraphStyle

    cell_style = ParagraphStyle(
        "PDF_Cell_Normal",
        parent=styles["BodyText"],
        fontName="PDF_Font",
        fontSize=6.5,
        leading=7.2,
        textColor=colors.black,
    )

    cell_style_bold = ParagraphStyle(
        "PDF_Cell_Bold",
        parent=cell_style,
        fontName="PDF_Font_Bold",
        fontSize=7.5,
        leading=8.0,
        textColor=colors.black,
    )

    cell_style_white = ParagraphStyle(
        "PDF_Cell_White",
        parent=cell_style,
        fontName="PDF_Font_Bold",
        fontSize=8.5,
        leading=9.0,
        textColor=colors.white,
    )

    elements = [Paragraph(title, styles["Title"]), Spacer(1, 4)]

    display_df = pdf_df.copy()
    for col in display_df.columns:
        if col not in ["AOP", "Pozicija", "Naziv", "Red", "Kategorija", "Opis", "Status"]:
            display_df[col] = display_df[col].apply(format_pdf_value)

    # Redovi koi treba da bidat potencirani vo PDF isto kako na ekran
    subtotal_rows = [
        "201", "207", "213", "223", "224", "234", "235",
        "001", "002", "009", "010", "020", "021", "031",
        "036", "037", "045", "052", "053", "059", "065",
        "071", "081", "082", "085", "095"
    ]

    main_total_rows = ["063", "111", "246", "247", "250", "251"]

    data = [list(display_df.columns)]
    row_aops = []

    for _, row in display_df.iterrows():
        aop = ""
        if "AOP" in display_df.columns:
            aop = str(row.get("AOP", "")).strip().replace(".0", "").zfill(3)
        row_aops.append(aop)

        if aop in main_total_rows:
            paragraph_style = cell_style_white
        elif aop in subtotal_rows:
            paragraph_style = cell_style_bold
        else:
            paragraph_style = cell_style

        pdf_row = []
        for col in display_df.columns:
            value = row[col]
            if col in ["Pozicija", "Naziv", "Opis"]:
                pdf_row.append(Paragraph(str(value), paragraph_style))
            else:
                pdf_row.append(str(value))
        data.append(pdf_row)

    col_count = len(display_df.columns)
    if col_count == 3:
        col_widths = [42, 630, 105]
    elif col_count == 4:
        col_widths = [38, 500, 115, 115]
    else:
        col_widths = None

    table = Table(data, repeatRows=1, colWidths=col_widths)

    style = TableStyle([
        # HEADER
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'PDF_Font_Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
        ('TOPPADDING', (0, 0), (-1, 0), 3),

        # BODY
        ('FONTNAME', (0, 1), (-1, -1), 'PDF_Font'),
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
        ('LEADING', (0, 1), (-1, -1), 7),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),

        # ALIGNMENTS
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # GRID
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor("#334155")),

        # PADDING
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 1), (-1, -1), 1.5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 1.5),
    ])

    for i, aop in enumerate(row_aops, start=1):
        if aop in subtotal_rows:
            style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#E5E7EB"))
            style.add("FONTNAME", (0, i), (-1, i), "PDF_Font_Bold")
            style.add("FONTSIZE", (0, i), (-1, i), 7.5)
            style.add("TEXTCOLOR", (0, i), (-1, i), colors.black)

        if aop in main_total_rows:
            style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#1E3A8A"))
            style.add("TEXTCOLOR", (0, i), (-1, i), colors.white)
            style.add("FONTNAME", (0, i), (-1, i), "PDF_Font_Bold")
            style.add("FONTSIZE", (0, i), (-1, i), 8.5)

    table.setStyle(style)
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output.getvalue()

def export_cash_flow_pdf(df):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = "PDF_Font_Bold"
    elements = [Paragraph("Извештај за паричните текови", styles["Title"]), Spacer(1, 14)]
    data = [list(df.columns)] + df.astype(str).values.tolist()
    table = Table(data, repeatRows=1, colWidths=[50, 420, 120])
    total_rows = ["CF08", "CF12", "CF16", "CF17", "CF20"]
    style = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B1F4D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "PDF_Font_Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "PDF_Font"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
    ])
    for i, row in enumerate(df.itertuples(index=False), start=1):
        aop = str(row.AOP)
        if aop in total_rows:
            style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#D9EAD3"))
            style.add("FONTNAME", (0, i), (-1, i), "PDF_Font_Bold")
            style.add("FONTSIZE", (0, i), (-1, i), 10)
        if aop == "CF17":
            style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#2F65D9"))
            style.add("TEXTCOLOR", (0, i), (-1, i), colors.white)
            style.add("FONTNAME", (0, i), (-1, i), "PDF_Font_Bold")
            style.add("FONTSIZE", (0, i), (-1, i), 11)
    table.setStyle(style)
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output.getvalue()




# =====================================================
# SIGMA STABLE ENGINE OVERRIDES - PDF/EXCEL + KLASI + DIJAGNOSTIKA
# =====================================================
import re
import pandas as pd
import numpy as np
import pdfplumber

REQUIRED_ZAKLUCEN_COLUMNS = [
    "prethodna_dolzi", "prethodna_pobaruva",
    "tekovna_dolzi", "tekovna_pobaruva",
    "vkupno_dolzi", "vkupno_pobaruva",
    "krajno_dolzi", "krajno_pobaruva",
]

KONTO_CLASS_NAMES = {
    "0": "Класа 0 - Нетековни средства",
    "1": "Класа 1 - Пари, побарувања и краткорочни средства",
    "2": "Класа 2 - Обврски",
    "3": "Класа 3 - Залихи и материјали",
    "4": "Класа 4 - Трошоци",
    "5": "Класа 5 - Слободна класа / останато",
    "6": "Класа 6 - Производство и залихи",
    "7": "Класа 7 - Приходи",
    "8": "Класа 8 - Финансиски резултат",
    "9": "Класа 9 - Капитал и резерви",
}


def normalize_konto(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    s = s.replace(".0", "")
    s = re.sub(r"\s+", "", s)
    s = s.replace("-", "").replace("/", "")
    if re.fullmatch(r"\d+", s) and len(s) < 3:
        s = s.zfill(3)
    return s


def normalize_rule_prefix(x):
    """
    Нормализација за правила/префикси.
    ВАЖНО: тука НЕ правиме zfill.
    Причина: правило 73,74 мора да ги фати 740011, 741U1, 742U;
    ако 74 стане 074, приходите нема да се читаат.
    """
    if pd.isna(x):
        return ""
    s = str(x).strip()
    s = s.replace(".0", "")
    s = re.sub(r"\s+", "", s)
    s = s.replace("-", "").replace("/", "")
    return s.upper()


def _numeric_parent_prefixes(prefix, min_len=3):
    """
    Враќа можни синтетички родители за аналитичко правило.

    Пример:
    - 2340 -> 234
    - 2341 -> 234
    - 2400 -> 240
    - 41400 -> 4140, 414, но ќе се земе првиот што постои во документот.

    Ова го решава случајот кога правилата се пишани аналитички,
    а конкретниот заклучен лист има само синтетички редови.
    """
    prefix = normalize_rule_prefix(prefix)
    if not prefix:
        return []

    # Fallback нагоре има смисла само за чисто нумерички конта.
    # Алфанумерички конта како 741U1 ги оставаме како што се.
    if not re.fullmatch(r"\d+", prefix):
        return []

    parents = []
    for length in range(len(prefix) - 1, min_len - 1, -1):
        parents.append(prefix[:length])

    return parents


def _find_existing_synthetic_parent(konta_norm, prefix, min_len=3):
    """
    Ако нема погодени аналитики за правило, бара најблизок постоечки
    синтетички родител во документот.

    Важно: бара EXACT parent, не startswith parent, за да не внесе други
    аналитики погрешно.
    """
    for parent in _numeric_parent_prefixes(prefix, min_len=min_len):
        parent_idxs = [idx for idx, konto in konta_norm.items() if str(konto) == parent]
        if parent_idxs:
            return parent_idxs, parent

    return [], ""


def _konto_sort_key(k):
    k = normalize_konto(k)
    m = re.match(r"^(\d+)", k)
    n = int(m.group(1)) if m else 999999
    return (n, len(k), k)


def _is_valid_konto(k):
    k = normalize_konto(k)
    return bool(re.match(r"^\d{1,6}[A-Za-zА-Ша-ш0-9]*$", k)) and len(k) >= 3


def _join_suffix_token(tok):
    tok = str(tok).strip()
    if not tok:
        return False
    if tok.startswith("+") or tok.startswith("-"):
        return False
    if len(tok) > 3:
        return False
    if any(ch.isdigit() for ch in tok):
        return True
    if tok.isalpha() and tok.upper() == tok and len(tok) <= 2:
        return True
    return False


def extract_pdf_konto_naziv(left_part):
    """
    Стабилно вадење конто од PDF line-text.
    Не го лепи описот кон конто: пример '102 sredstva...' останува конто 102,
    а формати како '310 1', '741 U1', '742 GP1' се нормализираат во 3101, 741U1, 742GP1.
    """
    left_part = str(left_part).strip()
    if not left_part:
        return None, None

    parts = left_part.split()
    if not parts:
        return None, None

    first = parts[0].strip()
    if not re.fullmatch(r"\d{3,6}[A-Za-zА-Ша-ш0-9]*", first):
        return None, None

    konto = first
    desc_start = 1
    if len(parts) > 1 and _join_suffix_token(parts[1]):
        konto = konto + parts[1]
        desc_start = 2

    konto = normalize_konto(konto)
    if not _is_valid_konto(konto):
        return None, None

    naziv = " ".join(parts[desc_start:]).strip()
    return konto, naziv


def _pdf_number_pattern():
    return r"-?(?:\d{1,3}(?:[\.,]\d{3})+[\.,]\d{2}|\d+[\.,]\d{2})"


def finalize_zaklucen_dataframe(df):
    """Финално чистење, нормализација и групирање на прочитан заклучен лист."""
    if df is None or df.empty:
        return pd.DataFrame(columns=["konto", "naziv"] + REQUIRED_ZAKLUCEN_COLUMNS)

    df = df.copy()
    if "konto" not in df.columns:
        return pd.DataFrame(columns=["konto", "naziv"] + REQUIRED_ZAKLUCEN_COLUMNS)
    if "naziv" not in df.columns:
        df["naziv"] = ""

    df["konto"] = df["konto"].apply(normalize_konto)
    df = df[df["konto"].apply(_is_valid_konto)].copy()

    for col in REQUIRED_ZAKLUCEN_COLUMNS:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = df[col].apply(clean_number)

    df["naziv"] = df["naziv"].fillna("").astype(str).str.strip()
    numeric_sum = df[REQUIRED_ZAKLUCEN_COLUMNS].abs().sum(axis=1)
    df = df[numeric_sum > 0].copy()

    if df.empty:
        return pd.DataFrame(columns=["konto", "naziv"] + REQUIRED_ZAKLUCEN_COLUMNS)

    # Ако истиот parser фати исто конто повеќе пати, собери ги бројките.
    # Првиот најдолг опис го задржуваме за приказ.
    agg = {col: "sum" for col in REQUIRED_ZAKLUCEN_COLUMNS}
    desc = (
        df.sort_values("naziv", key=lambda s: s.str.len(), ascending=False)
          .drop_duplicates("konto")
          .set_index("konto")["naziv"]
    )
    out = df.groupby("konto", as_index=False).agg(agg)
    out["naziv"] = out["konto"].map(desc).fillna("")
    out = out[["konto", "naziv"] + REQUIRED_ZAKLUCEN_COLUMNS]
    out = out.sort_values("konto", key=lambda s: s.map(_konto_sort_key)).reset_index(drop=True)
    return out


def read_zaklucen_pdf_standard_lines(pdf):
    rows = []
    number_pattern = _pdf_number_pattern()
    for page in pdf.pages:
        text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
        for line in text.split("\n"):
            line = line.strip()
            nums = re.findall(number_pattern, line)
            if not (len(nums) >= 8 or len(nums) == 6):
                continue
            first_num = re.search(number_pattern, line)
            if not first_num:
                continue
            left_part = line[:first_num.start()].strip()
            konto, naziv = extract_pdf_konto_naziv(left_part)
            if not konto:
                continue
            if len(nums) >= 8:
                vals = [pdf_number(x) for x in nums[:8]]
                prethodna_dolzi, prethodna_pobaruva, tekovna_dolzi, tekovna_pobaruva, vkupno_dolzi, vkupno_pobaruva, krajno_dolzi, krajno_pobaruva = vals
            else:
                prethodna_dolzi, prethodna_pobaruva = split_signed_saldo(nums[0])
                tekovna_dolzi = pdf_number(nums[1])
                tekovna_pobaruva = pdf_number(nums[2])
                vkupno_dolzi = pdf_number(nums[3])
                vkupno_pobaruva = pdf_number(nums[4])
                krajno_dolzi, krajno_pobaruva = split_signed_saldo(nums[5])
            rows.append({
                "konto": konto,
                "naziv": naziv,
                "prethodna_dolzi": prethodna_dolzi,
                "prethodna_pobaruva": prethodna_pobaruva,
                "tekovna_dolzi": tekovna_dolzi,
                "tekovna_pobaruva": tekovna_pobaruva,
                "vkupno_dolzi": vkupno_dolzi,
                "vkupno_pobaruva": vkupno_pobaruva,
                "krajno_dolzi": krajno_dolzi,
                "krajno_pobaruva": krajno_pobaruva,
            })
    return finalize_zaklucen_dataframe(pd.DataFrame(rows))


def _pdf_balance_diff(df):
    if df is None or df.empty:
        return float("inf")
    if not all(c in df.columns for c in REQUIRED_ZAKLUCEN_COLUMNS):
        return float("inf")
    diff = 0.0
    for d_col, p_col in [
        ("prethodna_dolzi", "prethodna_pobaruva"),
        ("tekovna_dolzi", "tekovna_pobaruva"),
        ("vkupno_dolzi", "vkupno_pobaruva"),
        ("krajno_dolzi", "krajno_pobaruva"),
    ]:
        diff += abs(round(float(df[d_col].sum() - df[p_col].sum()), 2))
    return diff


def _parser_quality_metrics(df):
    if df is None or df.empty:
        return {"rows": 0, "diff": float("inf"), "valid": 0, "has_revenue": 0, "has_expense": 0}
    return {
        "rows": int(len(df)),
        "diff": _pdf_balance_diff(df),
        "valid": int(df["konto"].apply(_is_valid_konto).sum()) if "konto" in df.columns else 0,
        "has_revenue": int(df["konto"].astype(str).str.startswith("7").any()) if "konto" in df.columns else 0,
        "has_expense": int(df["konto"].astype(str).str.startswith("4").any()) if "konto" in df.columns else 0,
    }


def _pdf_parser_score(df, priority=0):
    m = _parser_quality_metrics(df)
    if m["rows"] <= 5:
        return (-1, -999999, -999999, -priority)
    balanced = 1 if m["diff"] <= 1.0 else 0
    return (
        balanced,
        m["has_revenue"] + m["has_expense"],
        -m["diff"],
        m["valid"],
        m["rows"],
        -priority,
    )


def _standard_pdf_from_page_texts(page_texts):
    """STANDARD/ZONEL parser from text already extracted once per page."""
    rows = []
    number_pattern = _pdf_number_pattern()

    for text in page_texts:
        for line in (text or "").split("\n"):
            line = line.strip()
            nums = re.findall(number_pattern, line)
            if not (len(nums) >= 8 or len(nums) == 6):
                continue

            first_num = re.search(number_pattern, line)
            if not first_num:
                continue

            left_part = line[:first_num.start()].strip()
            konto, naziv = extract_pdf_konto_naziv(left_part)
            if not konto:
                continue

            if len(nums) >= 8:
                vals = [pdf_number(x) for x in nums[:8]]
                (
                    prethodna_dolzi, prethodna_pobaruva,
                    tekovna_dolzi, tekovna_pobaruva,
                    vkupno_dolzi, vkupno_pobaruva,
                    krajno_dolzi, krajno_pobaruva,
                ) = vals
            else:
                prethodna_dolzi, prethodna_pobaruva = split_signed_saldo(nums[0])
                tekovna_dolzi = pdf_number(nums[1])
                tekovna_pobaruva = pdf_number(nums[2])
                vkupno_dolzi = pdf_number(nums[3])
                vkupno_pobaruva = pdf_number(nums[4])
                krajno_dolzi, krajno_pobaruva = split_signed_saldo(nums[5])

            rows.append({
                "konto": konto,
                "naziv": naziv,
                "prethodna_dolzi": prethodna_dolzi,
                "prethodna_pobaruva": prethodna_pobaruva,
                "tekovna_dolzi": tekovna_dolzi,
                "tekovna_pobaruva": tekovna_pobaruva,
                "vkupno_dolzi": vkupno_dolzi,
                "vkupno_pobaruva": vkupno_pobaruva,
                "krajno_dolzi": krajno_dolzi,
                "krajno_pobaruva": krajno_pobaruva,
            })

    return finalize_zaklucen_dataframe(pd.DataFrame(rows))


def _pdf_safe_print(message):
    """Render diagnostic marker; flush immediately so the last completed phase is visible."""
    try:
        print(f"[PDF SAFE V2] {message}", flush=True)
    except Exception:
        pass


def read_zaklucen_pdf(file):
    """
    Render-safe PDF router.

    Key rule: extract page text only ONCE. STANDARD/ZONEL is parsed from that
    already-extracted text. Only when STANDARD is not reliable do we run ONE
    positional fallback parser (X-position, CFMA or IVA), never all of them.
    """
    try:
        st.success("PDF успешно е вчитан. Се врши безбедно препознавање на структурата...")
    except Exception:
        pass

    # Keep independent bytes so any fallback can reopen the PDF from position 0.
    try:
        if hasattr(file, "getvalue"):
            pdf_bytes = file.getvalue()
        else:
            pos = file.tell() if hasattr(file, "tell") else None
            if hasattr(file, "seek"):
                file.seek(0)
            pdf_bytes = file.read()
            if pos is not None and hasattr(file, "seek"):
                file.seek(pos)
    except Exception:
        if hasattr(file, "seek"):
            file.seek(0)
        pdf_bytes = file.read()

    _pdf_safe_print(f"START bytes={len(pdf_bytes):,}")

    page_texts = []
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            _pdf_safe_print(f"OPEN OK pages={len(pdf.pages)}")
            for page_no, page in enumerate(pdf.pages, start=1):
                # One and only text extraction pass for routing + STANDARD parser.
                text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
                page_texts.append(text)
                _pdf_safe_print(f"TEXT page={page_no} chars={len(text)}")
    except Exception as e:
        _pdf_safe_print(f"TEXT ERROR {type(e).__name__}: {e}")
        try:
            st.error(f"PDF не може да се прочита: {e}")
            st.stop()
        except Exception:
            pass
        return pd.DataFrame(columns=["konto", "naziv"] + REQUIRED_ZAKLUCEN_COLUMNS)

    text_all = "\n".join(page_texts)
    text_upper = text_all.upper()

    # 1) Fast path: STANDARD/ZONEL from the same text already extracted above.
    try:
        df_standard = _standard_pdf_from_page_texts(page_texts)
        m_standard = _parser_quality_metrics(df_standard)
        _pdf_safe_print(
            f"STANDARD rows={m_standard['rows']} diff={m_standard['diff']:.2f} "
            f"rev={m_standard['has_revenue']} exp={m_standard['has_expense']}"
        )
    except Exception as e:
        _pdf_safe_print(f"STANDARD ERROR {type(e).__name__}: {e}")
        df_standard = pd.DataFrame(columns=["konto", "naziv"] + REQUIRED_ZAKLUCEN_COLUMNS)
        m_standard = _parser_quality_metrics(df_standard)

    # A balanced standard parse with enough rows is final. No extract_words pass.
    if m_standard["rows"] > 5 and m_standard["diff"] <= 1.0:
        _pdf_safe_print("ROUTE=STANDARD short-circuit; positional parsers skipped")
        try:
            st.info(
                f"Детектиран PDF формат: STANDARD / ZONEL / line parser "
                f"(редови: {len(df_standard)}, контролна разлика: {_pdf_balance_diff(df_standard):,.2f})."
            )
        except Exception:
            pass
        return finalize_zaklucen_dataframe(df_standard)

    # 2) Choose at most ONE positional fallback based on document signature.
    fallback_name = None
    fallback_func = None

    # IVA signature is the most specific; check before generic BRUTO BILANS.
    if "БРУТОБИЛАНС" in text_upper or "BBILANS" in text_upper:
        fallback_name = "IVA SOFT / BBILANS"
        fallback_func = read_pdf_ivasoft_format
    elif "BRUTO" in text_upper and "BILANS" in text_upper:
        fallback_name = "CFMA / Info Biro"
        fallback_func = read_pdf_cfma_infobiro_format
    elif "ZAKLU" in text_upper and "SALDO" in text_upper and "TEKOV" in text_upper:
        fallback_name = "X-position parser"
        fallback_func = read_zaklucen_pdf_words_format

    candidates = []
    if m_standard["rows"] > 5:
        candidates.append({"name": "STANDARD / ZONEL / line parser", "df": df_standard, "priority": 1})

    if fallback_func is not None:
        _pdf_safe_print(f"FALLBACK START {fallback_name}")
        try:
            with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
                df_fallback = finalize_zaklucen_dataframe(fallback_func(pdf))
            m_fb = _parser_quality_metrics(df_fallback)
            _pdf_safe_print(f"FALLBACK END {fallback_name} rows={m_fb['rows']} diff={m_fb['diff']:.2f}")
            if m_fb["rows"] > 5:
                candidates.append({"name": fallback_name, "df": df_fallback, "priority": 2})
        except Exception as e:
            _pdf_safe_print(f"FALLBACK ERROR {fallback_name}: {type(e).__name__}: {e}")
    else:
        _pdf_safe_print("FALLBACK=NONE by signature")

    if candidates:
        best = sorted(
            candidates,
            key=lambda c: _pdf_parser_score(c["df"], c["priority"]),
            reverse=True,
        )[0]
        best_df = finalize_zaklucen_dataframe(best["df"])
        _pdf_safe_print(f"ROUTE={best['name']} rows={len(best_df)} diff={_pdf_balance_diff(best_df):.2f}")
        try:
            st.info(
                f"Детектиран PDF формат: {best['name']} "
                f"(редови: {len(best_df)}, контролна разлика: {_pdf_balance_diff(best_df):,.2f})."
            )
        except Exception:
            pass
        return best_df

    _pdf_safe_print("NO VALID PARSER RESULT")
    try:
        st.error("PDF е прочитан, но не се пронајдени валидни редови од заклучниот лист / бруто билансот.")
        st.stop()
    except Exception:
        pass
    return pd.DataFrame(columns=["konto", "naziv"] + REQUIRED_ZAKLUCEN_COLUMNS)

def read_zaklucen(file):
    engine = "xlrd" if file.name.lower().endswith(".xls") else "openpyxl"
    raw = read_excel_fill_merged(file, engine=engine, sheet_name=0)

    start_row = None
    for i in range(len(raw)):
        row_values = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
        first_val = row_values[0] if row_values else ""
        if first_val in ["konto", "конто"] or re.fullmatch(r"\d{3,6}", first_val or ""):
            start_row = i
            break

    if start_row is None:
        try:
            st.error("Не може да се најде почеток на заклучниот лист.")
            st.stop()
        except Exception:
            pass
        return pd.DataFrame(columns=["konto", "naziv"] + REQUIRED_ZAKLUCEN_COLUMNS)

    first_val = str(raw.iloc[start_row, 0]).strip().lower()
    if first_val in ["konto", "конто"]:
        start_row += 1

    df = raw.iloc[start_row:].copy()
    df = df.dropna(how="all")

    if df.shape[1] >= 10:
        df = df.iloc[:, :10]
        df.columns = ["konto", "naziv"] + REQUIRED_ZAKLUCEN_COLUMNS
    else:
        df = df.iloc[:, :9]
        df.columns = [
            "konto", "naziv", "saldo",
            "prethodna_dolzi", "prethodna_pobaruva",
            "tekovna_dolzi", "tekovna_pobaruva",
            "vkupno_dolzi", "vkupno_pobaruva",
        ]
        df["saldo"] = df["saldo"].apply(clean_number)
        df["krajno_dolzi"] = df["saldo"].apply(lambda x: x if x > 0 else 0)
        df["krajno_pobaruva"] = df["saldo"].apply(lambda x: abs(x) if x < 0 else 0)

    return finalize_zaklucen_dataframe(df)


def _matching_nonoverlap_indices(df, prefixes):
    if df is None or df.empty or "konto" not in df.columns:
        return []

    prefixes = [normalize_rule_prefix(p) for p in prefixes]
    prefixes = [p for p in prefixes if p]
    if not prefixes:
        return []

    konta = df["konto"].astype(str).apply(normalize_konto)
    selected = []
    selected_codes = []

    for idx, konto in sorted(konta.items(), key=lambda kv: _konto_sort_key(kv[1])):
        # Правилата се префикси: 74 мора да фати 740011, 741U1, 742U итн.
        if not any(konto.startswith(prefix) for prefix in prefixes):
            continue

        # Ако веќе има избрано пократко синтетичко конто кое го покрива ова,
        # не го земаме повторно за да нема дуплирање синтетика + аналитика.
        if any(konto != parent and konto.startswith(parent) for parent in selected_codes):
            continue

        selected.append(idx)
        selected_codes.append(konto)

    return selected


def sum_rule(df, konto_text, kolona, operacija):
    """
    Стабилно сумирање по конта без дуплирање синтетички и аналитички редови.
    Ако постои 400 и 40001, за правило 400 се зема 400, а не 400 + 40001.
    Ако не постои синтетика, се земаат аналитиките.
    """
    if pd.isna(konto_text) or pd.isna(kolona):
        return 0.0
    kolona = str(kolona).strip()
    if kolona not in df.columns:
        return 0.0
    # За правила не смее да се користи normalize_konto(), бидејќи тоа би го претворило
    # 74 во 074 и нема да ги фати приходните конта 740/741/742.
    prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
    prefixes = [x for x in prefixes if x]
    if not prefixes:
        return 0.0
    idxs = _matching_nonoverlap_indices(df, prefixes)
    if not idxs:
        return 0.0
    value = float(df.loc[idxs, kolona].apply(clean_number).sum())
    if str(operacija).strip() == "-":
        return -value
    return value


def build_konto_class_summary(df, mode="raw"):
    """
    Збир по класи.
    mode="raw" ги прикажува сите прочитани редови и затоа мора да се сложува со контролата на документот.
    mode="rules" користи не-дуплирачка селекција како sum_rule и служи за дијагностика на правила.
    """
    df = finalize_zaklucen_dataframe(df)
    rows = []
    for cls in [str(i) for i in range(10)]:
        if mode == "rules":
            idxs = _matching_nonoverlap_indices(df, [cls])
            part = df.loc[idxs].copy() if idxs else pd.DataFrame(columns=df.columns)
        else:
            part = df[df["konto"].astype(str).str.startswith(cls, na=False)].copy()
        row = {
            "Класа": cls,
            "Назив": KONTO_CLASS_NAMES.get(cls, f"Класа {cls}"),
            "Број на редови": int(len(part)),
        }
        for col in REQUIRED_ZAKLUCEN_COLUMNS:
            row[col] = float(part[col].sum()) if col in part.columns else 0.0
        row["Нето тековна"] = row["tekovna_dolzi"] - row["tekovna_pobaruva"]
        row["Нето крајно"] = row["krajno_dolzi"] - row["krajno_pobaruva"]
        rows.append(row)

    out = pd.DataFrame(rows)
    total = {"Класа": "Вкупно", "Назив": "Контрола по класи", "Број на редови": int(out["Број на редови"].sum())}
    for col in REQUIRED_ZAKLUCEN_COLUMNS + ["Нето тековна", "Нето крајно"]:
        total[col] = float(out[col].sum())
    out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)
    return out


def validate_zaklucen_list(df):
    df = finalize_zaklucen_dataframe(df)
    checks = []
    pairs = [
        ("Prethodna godina", "prethodna_dolzi", "prethodna_pobaruva"),
        ("Tekovna godina", "tekovna_dolzi", "tekovna_pobaruva"),
        ("Vkupno", "vkupno_dolzi", "vkupno_pobaruva"),
        ("Krajno saldo", "krajno_dolzi", "krajno_pobaruva"),
    ]
    for label, debit_col, credit_col in pairs:
        debit = float(df[debit_col].sum()) if debit_col in df.columns else 0.0
        credit = float(df[credit_col].sum()) if credit_col in df.columns else 0.0
        diff = round(debit - credit, 2)
        checks.append({
            "Kontrola": label,
            "Dolzi": round(debit, 2),
            "Pobaruva": round(credit, 2),
            "Razlika": diff,
            "Status": "✅ OK" if abs(diff) <= 1 else "❌ Razlika",
        })
    return pd.DataFrame(checks)


def _aop_value(bs, aop, col):
    if bs is None or bs.empty or col not in bs.columns:
        return 0.0
    a = str(aop).zfill(3)
    m = bs["AOP"].astype(str).str.replace(".0", "", regex=False).str.zfill(3) == a
    return float(bs.loc[m, col].sum()) if m.any() else 0.0


def _bu_value(bu, aop):
    if bu is None or bu.empty or "Iznos" not in bu.columns:
        return 0.0
    a = str(aop).zfill(3)
    m = bu["AOP"].astype(str).str.replace(".0", "", regex=False).str.zfill(3) == a
    return float(bu.loc[m, "Iznos"].sum()) if m.any() else 0.0


def build_bs_class_diagnostic(df, bu, bs):
    class_df = build_konto_class_summary(df)
    class_map = {str(r["Класа"]): r for _, r in class_df[class_df["Класа"] != "Вкупно"].iterrows()}

    net_result_bu = _bu_value(bu, "246") - _bu_value(bu, "247")
    aktiva = _aop_value(bs, "063", "Tekovna godina")
    pasiva = _aop_value(bs, "111", "Tekovna godina")
    aktiva_prev = _aop_value(bs, "063", "Prethodna godina")
    pasiva_prev = _aop_value(bs, "111", "Prethodna godina")

    # Индикативни класи, не е официјален биланс - служи за насочување каде да се бара разликата.
    assets_net = sum(float(class_map.get(c, {}).get("Нето крајно", 0.0)) for c in ["0", "1", "3", "6"])
    liabilities_equity_net = -sum(float(class_map.get(c, {}).get("Нето крајно", 0.0)) for c in ["2", "9"])
    income_expense_net = -sum(float(class_map.get(c, {}).get("Нето тековна", 0.0)) for c in ["4", "7", "8"])

    rows = [
        {
            "Контрола": "БС тековна година - Актива / Пасива",
            "Износ 1": aktiva,
            "Износ 2": pasiva,
            "Разлика": round(aktiva - pasiva, 2),
            "Коментар": "Ова е официјалната контрола по AOP 063 и AOP 111.",
        },
        {
            "Контрола": "БС претходна година - Актива / Пасива",
            "Износ 1": aktiva_prev,
            "Износ 2": pasiva_prev,
            "Разлика": round(aktiva_prev - pasiva_prev, 2),
            "Коментар": "Контрола за претходна година по AOP 063 и AOP 111.",
        },
        {
            "Контрола": "Класи 0+1+3+6 - индикативна актива",
            "Износ 1": assets_net,
            "Износ 2": aktiva,
            "Разлика": round(assets_net - aktiva, 2),
            "Коментар": "Помага да се види дали активните класи се согласуваат со БС активата.",
        },
        {
            "Контрола": "Класи 2+9 - индикативна пасива/капитал",
            "Износ 1": liabilities_equity_net,
            "Износ 2": pasiva,
            "Разлика": round(liabilities_equity_net - pasiva, 2),
            "Коментар": "Помага да се види дали обврските и капиталот се согласуваат со БС пасивата.",
        },
        {
            "Контрола": "БУ резултат 246-247 vs класи 4/7/8",
            "Износ 1": net_result_bu,
            "Износ 2": income_expense_net,
            "Разлика": round(net_result_bu - income_expense_net, 2),
            "Коментар": "Ако има разлика, проверете правила за приходи/расходи или дуплирање синтетика/аналитика.",
        },
    ]
    out = pd.DataFrame(rows)
    out["Статус"] = out["Разлика"].apply(lambda x: "✅ OK" if abs(float(x)) <= 1 else "⚠️ Проверка")
    return out


def build_rule_trace(df, rules_file, sheet_name, year_mode="current"):
    """
    Дијагностика за трансфер од заклучен лист во BU/BS.
    За секое правило покажува кои конта се погодени и колкав износ влегол.
    year_mode="current" ги чита колоните konto/kolona/Operacija.
    year_mode="previous" ги чита konto_prethodna/kolona_prethodna/Operacija_prethodna ако постојат.
    """
    try:
        rules = read_rules(rules_file, sheet_name)
    except Exception as e:
        return pd.DataFrame([{
            "Sheet": sheet_name,
            "Грешка": str(e)
        }])

    rows = []
    for _, row in rules.iterrows():
        aop = format_aop(row.get("AOP", ""))
        pozicija = str(row.get("Pozicija", "")).strip()

        if year_mode == "previous":
            konto_text = row.get("konto_prethodna", row.get("konto", None))
            kolona = row.get("kolona_prethodna", None)
            operacija = row.get("Operacija_prethodna", row.get("Operacija", "+"))
        else:
            konto_text = row.get("konto", None)
            kolona = row.get("kolona", None)
            operacija = row.get("Operacija", "+")

        if pd.isna(konto_text) or pd.isna(kolona):
            continue

        prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
        prefixes = [x for x in prefixes if x]
        idxs = _matching_nonoverlap_indices(df, prefixes) if prefixes else []

        if idxs:
            matched_df = df.loc[idxs].copy()
            matched_konta = matched_df["konto"].astype(str).tolist()
            value = float(matched_df[str(kolona).strip()].apply(clean_number).sum())
        else:
            matched_konta = []
            value = 0.0

        if str(operacija).strip() == "-":
            value = -value

        rows.append({
            "AOP": aop,
            "Позиција": pozicija,
            "Правило конта": str(konto_text),
            "Колона": str(kolona),
            "Операција": str(operacija),
            "Број конта": len(matched_konta),
            "Погодени конта": ", ".join(matched_konta[:30]) + (" ..." if len(matched_konta) > 30 else ""),
            "Износ": value,
            "Статус": "✅ OK" if len(matched_konta) > 0 else "⚠️ Нема погодени конта",
        })

    return pd.DataFrame(rows)




# =====================================================
# SIGMA v4 BU PREFIX ENGINE OVERRIDES
# =====================================================

def _matching_prefix_all_indices(df, prefixes):
    """
    За BU (приходи/расходи) правилата мора да работат како вистински filter startswith.
    Пример: правило 400 треба да ги собере 400, 40001, 40010, 40013...
    Ова важи затоа што во трошоци/приходи често редовите не се синтетика + аналитика,
    туку посебни конта кои сите треба да влезат во БУ.
    """
    if df is None or df.empty or "konto" not in df.columns:
        return []

    prefixes = [normalize_rule_prefix(p) for p in prefixes]
    prefixes = [p for p in prefixes if p]
    if not prefixes:
        return []

    konta = df["konto"].astype(str).apply(normalize_konto)
    mask = konta.apply(lambda k: any(str(k).startswith(prefix) for prefix in prefixes))
    return list(df.index[mask])


def sum_rule_prefix_all(df, konto_text, kolona, operacija):
    """
    BU сумирање: земи СИТЕ конта што почнуваат со правилото.
    Ова ја решава ситуацијата 400 + 40001, 414 + 41400, 740011 + 740021 итн.
    """
    if pd.isna(konto_text) or pd.isna(kolona):
        return 0.0

    kolona = str(kolona).strip()
    if kolona not in df.columns:
        return 0.0

    prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
    prefixes = [x for x in prefixes if x]
    if not prefixes:
        return 0.0

    idxs = _matching_prefix_all_indices(df, prefixes)
    if not idxs:
        return 0.0

    value = float(df.loc[idxs, kolona].apply(clean_number).sum())
    if str(operacija).strip() == "-":
        return -value
    return value


def calculate_bilans_uspeh(df, rules_file):
    """
    v4: BU користи prefix-all matching.
    БС останува со non-overlap логика преку sum_rule(), за да не се наруши активата/пасивата.
    """
    rules = read_rules(rules_file, "Pravila Bilans Uspeh")
    values, positions, order = {}, {}, []

    for _, row in rules.iterrows():
        aop = format_aop(row["AOP"])
        if aop not in order:
            order.append(aop)

        positions[aop] = row.get("Pozicija", "")

        val = sum_rule_prefix_all(
            df,
            row.get("konto", None),
            row.get("kolona", None),
            row.get("Operacija", "+")
        )
        values[aop] = values.get(aop, 0.0) + val

    values = apply_formulas_and_logic(rules, values)

    return pd.DataFrame([
        {
            "AOP": aop,
            "Pozicija": positions.get(aop, ""),
            "Iznos": values.get(aop, 0.0)
        }
        for aop in order
    ])


def build_rule_trace(df, rules_file, sheet_name, year_mode="current"):
    """
    Дијагностика за трансфер од заклучен лист во BU/BS.
    v4: за BU се користи prefix-all; за BS останува non-overlap.
    """
    try:
        rules = read_rules(rules_file, sheet_name)
    except Exception as e:
        return pd.DataFrame([{
            "Sheet": sheet_name,
            "Грешка": str(e)
        }])

    rows = []
    is_bu_sheet = str(sheet_name).strip().lower() == "pravila bilans uspeh".lower()

    for _, row in rules.iterrows():
        aop = format_aop(row.get("AOP", ""))
        pozicija = str(row.get("Pozicija", "")).strip()

        if year_mode == "previous":
            konto_text = row.get("konto_prethodna", row.get("konto", None))
            kolona = row.get("kolona_prethodna", None)
            operacija = row.get("Operacija_prethodna", row.get("Operacija", "+"))
        else:
            konto_text = row.get("konto", None)
            kolona = row.get("kolona", None)
            operacija = row.get("Operacija", "+")

        if pd.isna(konto_text) or pd.isna(kolona):
            continue

        prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
        prefixes = [x for x in prefixes if x]

        if prefixes:
            if is_bu_sheet:
                idxs = _matching_prefix_all_indices(df, prefixes)
                match_mode = "prefix-all"
            else:
                idxs = _matching_nonoverlap_indices(df, prefixes)
                match_mode = "non-overlap"
        else:
            idxs = []
            match_mode = "none"

        if idxs:
            matched_df = df.loc[idxs].copy()
            matched_konta = matched_df["konto"].astype(str).tolist()
            value = float(matched_df[str(kolona).strip()].apply(clean_number).sum())
        else:
            matched_konta = []
            value = 0.0

        if str(operacija).strip() == "-":
            value = -value

        rows.append({
            "AOP": aop,
            "Позиција": pozicija,
            "Правило конта": str(konto_text),
            "Колона": str(kolona),
            "Операција": str(operacija),
            "Начин": match_mode,
            "Број конта": len(matched_konta),
            "Погодени конта": ", ".join(matched_konta[:40]) + (" ..." if len(matched_konta) > 40 else ""),
            "Износ": value,
            "Статус": "✅ OK" if len(matched_konta) > 0 else "⚠️ Нема погодени конта",
        })

    return pd.DataFrame(rows)



# =====================================================
# SIGMA v5 SMART PREFIX ENGINE FOR BS
# =====================================================

def _matching_smart_prefix_indices(df, prefixes, kolona=None, tolerance=1.0):
    """
    Smart prefix matching за BS и дијагностика.

    Зошто е потребно:
    - Некои заклучни листови имаат и синтетика и аналитика, но синтетиката НЕ е секогаш збир.
      Пример: 669 има 0 крајно салдо, а 6691 има вистинско салдо -> non-overlap го земаше 669
      и погрешно ја прескокнуваше 6691.
    - Ако синтетичката сметка навистина е збир на аналитиките, тогаш ја земаме само синтетиката
      за да нема дуплирање.
    - Ако синтетиката е 0, а аналитиките имаат вредност, ги земаме аналитиките.
    - Ако синтетиката и аналитиките не се еднакви, ги земаме сите редови, затоа што во тој формат
      најчесто се посебни сметки, а не subtotal.
    """
    if df is None or df.empty or "konto" not in df.columns:
        return []

    prefixes = [normalize_rule_prefix(p) for p in prefixes]
    prefixes = [p for p in prefixes if p]
    if not prefixes:
        return []

    konta_norm = df["konto"].astype(str).apply(normalize_konto)
    selected = []

    for prefix in prefixes:
        prefix = normalize_rule_prefix(prefix)
        if not prefix:
            continue

        match_mask = konta_norm.apply(lambda k: str(k).startswith(prefix))
        match_idxs = list(df.index[match_mask])
        if not match_idxs:
            continue

        exact_idxs = [idx for idx in match_idxs if konta_norm.loc[idx] == prefix]
        child_idxs = [idx for idx in match_idxs if konta_norm.loc[idx] != prefix]

        if not exact_idxs or not child_idxs or kolona is None or str(kolona).strip() not in df.columns:
            chosen = match_idxs
        else:
            col = str(kolona).strip()
            exact_total = float(df.loc[exact_idxs, col].apply(clean_number).sum())
            child_total = float(df.loc[child_idxs, col].apply(clean_number).sum())

            if abs(exact_total - child_total) <= tolerance and abs(child_total) > tolerance:
                # Синтетиката изгледа како вистински subtotal -> земи ја само неа.
                chosen = exact_idxs
            elif abs(exact_total) <= tolerance and abs(child_total) > tolerance:
                # Синтетиката е празна/нулта, аналитиките носат салдо -> земи аналитики.
                chosen = child_idxs
            else:
                # Синтетиката и аналитиките не се ист збир -> најверојатно се посебни сметки.
                chosen = match_idxs

        for idx in chosen:
            if idx not in selected:
                selected.append(idx)

    return selected


def sum_rule_smart_prefix(df, konto_text, kolona, operacija):
    """
    BS smart сумирање.
    За разлика од старата non-overlap логика, ова не ги прескокнува 6631/6691
    само затоа што постои 663/669.
    """
    if pd.isna(konto_text) or pd.isna(kolona):
        return 0.0

    kolona = str(kolona).strip()
    if kolona not in df.columns:
        return 0.0

    prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
    prefixes = [x for x in prefixes if x]
    if not prefixes:
        return 0.0

    idxs = _matching_smart_prefix_indices(df, prefixes, kolona=kolona)
    if not idxs:
        return 0.0

    value = float(df.loc[idxs, kolona].apply(clean_number).sum())
    if str(operacija).strip() == "-":
        return -value
    return value


def calculate_bilans_sostojba(df, rules_file, bu=None):
    """
    v5: BS користи smart-prefix matching.
    Ова ги решава случаите 663/6631 и 669/6691, без слепо дуплирање.
    """
    rules = read_rules(rules_file, "Pravila bilans Sostojba")
    current_values, previous_values, positions, order = {}, {}, {}, []

    bu_current_map = {}
    if bu is not None:
        for _, bu_row in bu.iterrows():
            bu_aop = str(bu_row["AOP"]).strip().replace(".0", "").zfill(3)
            bu_current_map[f"BU{bu_aop}"] = float(bu_row["Iznos"])

    for _, row in rules.iterrows():
        aop = format_aop(row["AOP"])
        if aop not in order:
            order.append(aop)
        positions[aop] = row.get("Pozicija", "")

        current_val = sum_rule_smart_prefix(
            df,
            row.get("konto", None),
            row.get("kolona", None),
            row.get("Operacija", "+")
        )

        previous_val = sum_rule_smart_prefix(
            df,
            row.get("konto_prethodna", row.get("konto", None)),
            row.get("kolona_prethodna", None),
            row.get("Operacija_prethodna", row.get("Operacija", "+"))
        )

        current_values[aop] = current_values.get(aop, 0.0) + current_val
        previous_values[aop] = previous_values.get(aop, 0.0) + previous_val

    if bu is not None:
        bu255 = bu_current_map.get("BU255", 0.0)
        bu256 = bu_current_map.get("BU256", 0.0)
        current_values["077"] = bu255
        current_values["078"] = bu256

    current_values = raspredeli_019_proporcionalno(
        df,
        current_values,
        "krajno_pobaruva",
        "krajno_dolzi"
    )

    current_values = apply_formulas_and_logic(rules, current_values, external_values=bu_current_map)

    if bu is not None:
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)
        current_values = apply_formulas_and_logic(rules, current_values, external_values=bu_current_map)
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)

    previous_values = raspredeli_019_proporcionalno(
        df,
        previous_values,
        "prethodna_pobaruva",
        "prethodna_dolzi"
    )

    previous_values = apply_formulas_and_logic(rules, previous_values)

    return pd.DataFrame([
        {
            "AOP": aop,
            "Pozicija": positions.get(aop, ""),
            "Tekovna godina": current_values.get(aop, 0.0),
            "Prethodna godina": previous_values.get(aop, 0.0)
        }
        for aop in order
    ])


def build_rule_trace(df, rules_file, sheet_name, year_mode="current"):
    """
    v5 дијагностика:
    - BU: prefix-all
    - BS: smart-prefix
    """
    try:
        rules = read_rules(rules_file, sheet_name)
    except Exception as e:
        return pd.DataFrame([{
            "Sheet": sheet_name,
            "Грешка": str(e)
        }])

    rows = []
    is_bu_sheet = str(sheet_name).strip().lower() == "pravila bilans uspeh".lower()

    for _, row in rules.iterrows():
        aop = format_aop(row.get("AOP", ""))
        pozicija = str(row.get("Pozicija", "")).strip()

        if year_mode == "previous":
            konto_text = row.get("konto_prethodna", row.get("konto", None))
            kolona = row.get("kolona_prethodna", None)
            operacija = row.get("Operacija_prethodna", row.get("Operacija", "+"))
        else:
            konto_text = row.get("konto", None)
            kolona = row.get("kolona", None)
            operacija = row.get("Operacija", "+")

        if pd.isna(konto_text) or pd.isna(kolona):
            continue

        prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
        prefixes = [x for x in prefixes if x]

        if prefixes:
            if is_bu_sheet:
                idxs = _matching_prefix_all_indices(df, prefixes)
                match_mode = "prefix-all"
            else:
                idxs = _matching_smart_prefix_indices(df, prefixes, kolona=str(kolona).strip())
                match_mode = "smart-prefix"
        else:
            idxs = []
            match_mode = "none"

        if idxs:
            matched_df = df.loc[idxs].copy()
            matched_konta = matched_df["konto"].astype(str).tolist()
            value = float(matched_df[str(kolona).strip()].apply(clean_number).sum())
        else:
            matched_konta = []
            value = 0.0

        if str(operacija).strip() == "-":
            value = -value

        rows.append({
            "AOP": aop,
            "Позиција": pozicija,
            "Правило конта": str(konto_text),
            "Колона": str(kolona),
            "Операција": str(operacija),
            "Начин": match_mode,
            "Број конта": len(matched_konta),
            "Погодени конта": ", ".join(matched_konta[:40]) + (" ..." if len(matched_konta) > 40 else ""),
            "Износ": value,
            "Статус": "✅ OK" if len(matched_konta) > 0 else "⚠️ Нема погодени конта",
        })

    return pd.DataFrame(rows)



# =====================================================
# SIGMA v6 EFFECTIVE ACCOUNT ENGINE
# =====================================================


def _immediate_child_indices(konta_norm, parent_idx, match_idxs):
    """
    Ги враќа непосредните аналитики под дадена синтетика.
    Пример: ако има 400, 40001 и 400010, за 400 непосредно дете е 40001,
    а 400010 е дете на 40001. Ова спречува дупло споредување на subtotal-и.
    """
    parent = str(konta_norm.loc[parent_idx])
    children = []

    for child_idx in match_idxs:
        if child_idx == parent_idx:
            continue

        child = str(konta_norm.loc[child_idx])

        if not child.startswith(parent) or child == parent:
            continue

        has_intermediate_parent = False

        for other_idx in match_idxs:
            if other_idx in [parent_idx, child_idx]:
                continue

            other = str(konta_norm.loc[other_idx])

            if (
                other.startswith(parent)
                and child.startswith(other)
                and other != parent
                and other != child
            ):
                has_intermediate_parent = True
                break

        if not has_intermediate_parent:
            children.append(child_idx)

    return children


def _matching_effective_indices(df, prefixes, kolona=None, tolerance=1.0):
    """
    Универзален engine за пренос на конта во BU/BS.

    Логика:
    1) Ако постои само синтетика -> ја зема синтетиката.
    2) Ако постојат синтетика + аналитики и синтетиката е subtotal
       на аналитиките -> ја исклучува синтетиката и ги зема аналитиките.
    3) Ако постојат синтетика + аналитики, но синтетиката НЕ е subtotal
       на аналитиките -> ги зема и синтетиката и аналитиките, затоа што
       најверојатно има директни книжења и на синтетичката сметка.
    4) Ако нема синтетика, а има аналитики -> ги зема аналитиките.

    Ова е стабилно за:
    - PDF/Excel заклучни листови со subtotal редови
    - IVA SOFT / CFMA формати
    - случаи како 400 + 40001 или 414 + 41400
    - случаи како 663 + 6631 и 669 + 6691
    """
    if df is None or df.empty or "konto" not in df.columns:
        return []

    prefixes = [normalize_rule_prefix(p) for p in prefixes]
    prefixes = [p for p in prefixes if p]

    if not prefixes:
        return []

    konta_norm = df["konto"].astype(str).apply(normalize_konto)
    selected = []

    col = str(kolona).strip() if kolona is not None and not pd.isna(kolona) else None
    has_col = bool(col and col in df.columns)

    for prefix in prefixes:
        match_mask = konta_norm.apply(lambda k: str(k).startswith(prefix))
        match_idxs = list(df.index[match_mask])

        if not match_idxs:
            # Fallback: правилото е аналитичко, а документот има само синтетика.
            # Пример: правило 2340/2341, а заклучниот лист има само 234;
            # правило 2400, а документот има само 240.
            fallback_idxs, fallback_parent = _find_existing_synthetic_parent(konta_norm, prefix)
            if fallback_idxs:
                for fallback_idx in fallback_idxs:
                    if fallback_idx not in selected:
                        selected.append(fallback_idx)
            continue

        # Одиме од пократки конта кон подолги, за прво да се анализираат синтетиките.
        match_idxs = sorted(match_idxs, key=lambda idx: (len(str(konta_norm.loc[idx])), str(konta_norm.loc[idx])))

        for idx in match_idxs:
            konto = str(konta_norm.loc[idx])
            child_idxs = _immediate_child_indices(konta_norm, idx, match_idxs)

            # Нема аналитика под ова конто -> земи го редот.
            if not child_idxs:
                if idx not in selected:
                    selected.append(idx)
                continue

            # Ако не знаеме колона, побезбедно е да не ја земаме синтетиката,
            # бидејќи има аналитики под неа. Ова најмногу важи за trace без сума.
            if not has_col:
                continue

            parent_value = float(df.loc[[idx], col].apply(clean_number).sum())
            children_value = float(df.loc[child_idxs, col].apply(clean_number).sum())

            parent_is_subtotal = (
                abs(parent_value - children_value) <= tolerance
                and abs(children_value) > tolerance
            )

            parent_empty_children_have_value = (
                abs(parent_value) <= tolerance
                and abs(children_value) > tolerance
            )

            if parent_is_subtotal or parent_empty_children_have_value:
                # Синтетиката е само збир / празен носител -> не ја земаме.
                # Аналитиките ќе бидат земени кога ќе дојде нивниот ред.
                continue

            # Синтетиката има сопствено книжење различно од аналитиките -> земи ја и неа.
            if idx not in selected:
                selected.append(idx)

    return selected


def _matching_effective_details(df, prefixes, kolona=None, tolerance=1.0):
    """Истата логика како _matching_effective_indices, но враќа и објаснување за trace."""
    if df is None or df.empty or "konto" not in df.columns:
        return [], [], []

    prefixes = [normalize_rule_prefix(p) for p in prefixes]
    prefixes = [p for p in prefixes if p]

    konta_norm = df["konto"].astype(str).apply(normalize_konto)
    selected = []
    excluded = []
    notes = []

    col = str(kolona).strip() if kolona is not None and not pd.isna(kolona) else None
    has_col = bool(col and col in df.columns)

    for prefix in prefixes:
        match_mask = konta_norm.apply(lambda k: str(k).startswith(prefix))
        match_idxs = list(df.index[match_mask])

        if not match_idxs:
            fallback_idxs, fallback_parent = _find_existing_synthetic_parent(konta_norm, prefix)
            if fallback_idxs:
                for fallback_idx in fallback_idxs:
                    if fallback_idx not in selected:
                        selected.append(fallback_idx)
                notes.append(
                    f"{prefix}: нема аналитика во документот -> земена синтетика {fallback_parent}"
                )
            else:
                notes.append(f"{prefix}: нема погодени конта")
            continue

        match_idxs = sorted(match_idxs, key=lambda idx: (len(str(konta_norm.loc[idx])), str(konta_norm.loc[idx])))

        for idx in match_idxs:
            konto = str(konta_norm.loc[idx])
            child_idxs = _immediate_child_indices(konta_norm, idx, match_idxs)

            if not child_idxs:
                if idx not in selected:
                    selected.append(idx)
                continue

            if not has_col:
                if idx not in excluded:
                    excluded.append(idx)
                notes.append(f"{konto}: има аналитика -> синтетиката е исклучена")
                continue

            parent_value = float(df.loc[[idx], col].apply(clean_number).sum())
            children_value = float(df.loc[child_idxs, col].apply(clean_number).sum())

            parent_is_subtotal = (
                abs(parent_value - children_value) <= tolerance
                and abs(children_value) > tolerance
            )
            parent_empty_children_have_value = (
                abs(parent_value) <= tolerance
                and abs(children_value) > tolerance
            )

            if parent_is_subtotal:
                if idx not in excluded:
                    excluded.append(idx)
                notes.append(f"{konto}: синтетика=subtotal -> земени аналитики")
                continue

            if parent_empty_children_have_value:
                if idx not in excluded:
                    excluded.append(idx)
                notes.append(f"{konto}: синтетика=0 -> земени аналитики")
                continue

            if idx not in selected:
                selected.append(idx)
            notes.append(f"{konto}: директно книжење + аналитики -> земени и двете")

    return selected, excluded, notes


def sum_rule_effective(df, konto_text, kolona, operacija):
    """Сумира правило со v6 effective-account логика."""
    if pd.isna(konto_text) or pd.isna(kolona):
        return 0.0

    kolona = str(kolona).strip()

    if kolona not in df.columns:
        return 0.0

    prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
    prefixes = [x for x in prefixes if x]

    if not prefixes:
        return 0.0

    idxs = _matching_effective_indices(df, prefixes, kolona=kolona)

    if not idxs:
        return 0.0

    value = float(df.loc[idxs, kolona].apply(clean_number).sum())

    if str(operacija).strip() == "-":
        return -value

    return value


# sum_rule се override-ира за сите места каде што се повикува старото име.
def sum_rule(df, konto_text, kolona, operacija):
    return sum_rule_effective(df, konto_text, kolona, operacija)


# BU повеќе не оди со слеп prefix-all, туку со effective matching.
def sum_rule_prefix_all(df, konto_text, kolona, operacija):
    return sum_rule_effective(df, konto_text, kolona, operacija)


# BS smart-prefix се заменува со истата централна effective логика.
def sum_rule_smart_prefix(df, konto_text, kolona, operacija):
    return sum_rule_effective(df, konto_text, kolona, operacija)


def calculate_bilans_uspeh(df, rules_file):
    """v7: BU користи central effective-account matching + synthetic fallback."""
    rules = read_rules(rules_file, "Pravila Bilans Uspeh")
    values, positions, order = {}, {}, []

    for _, row in rules.iterrows():
        aop = format_aop(row["AOP"])

        if aop not in order:
            order.append(aop)

        positions[aop] = row.get("Pozicija", "")

        val = sum_rule_effective(
            df,
            row.get("konto", None),
            row.get("kolona", None),
            row.get("Operacija", "+")
        )

        values[aop] = values.get(aop, 0.0) + val

    values = apply_formulas_and_logic(rules, values)

    return pd.DataFrame([
        {
            "AOP": aop,
            "Pozicija": positions.get(aop, ""),
            "Iznos": values.get(aop, 0.0)
        }
        for aop in order
    ])


def calculate_bilans_sostojba(df, rules_file, bu=None):
    """v7: BS користи central effective-account matching + synthetic fallback."""
    rules = read_rules(rules_file, "Pravila bilans Sostojba")
    current_values, previous_values, positions, order = {}, {}, {}, []

    bu_current_map = {}

    if bu is not None:
        for _, bu_row in bu.iterrows():
            bu_aop = str(bu_row["AOP"]).strip().replace(".0", "").zfill(3)
            bu_current_map[f"BU{bu_aop}"] = float(bu_row["Iznos"])

    for _, row in rules.iterrows():
        aop = format_aop(row["AOP"])

        if aop not in order:
            order.append(aop)

        positions[aop] = row.get("Pozicija", "")

        current_val = sum_rule_effective(
            df,
            row.get("konto", None),
            row.get("kolona", None),
            row.get("Operacija", "+")
        )

        previous_val = sum_rule_effective(
            df,
            row.get("konto_prethodna", row.get("konto", None)),
            row.get("kolona_prethodna", None),
            row.get("Operacija_prethodna", row.get("Operacija", "+"))
        )

        current_values[aop] = current_values.get(aop, 0.0) + current_val
        previous_values[aop] = previous_values.get(aop, 0.0) + previous_val

    if bu is not None:
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)

    current_values = raspredeli_019_proporcionalno(
        df,
        current_values,
        "krajno_pobaruva",
        "krajno_dolzi"
    )

    current_values = apply_formulas_and_logic(rules, current_values, external_values=bu_current_map)

    if bu is not None:
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)
        current_values = apply_formulas_and_logic(rules, current_values, external_values=bu_current_map)
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)

    previous_values = raspredeli_019_proporcionalno(
        df,
        previous_values,
        "prethodna_pobaruva",
        "prethodna_dolzi"
    )

    previous_values = apply_formulas_and_logic(rules, previous_values)

    return pd.DataFrame([
        {
            "AOP": aop,
            "Pozicija": positions.get(aop, ""),
            "Tekovna godina": current_values.get(aop, 0.0),
            "Prethodna godina": previous_values.get(aop, 0.0)
        }
        for aop in order
    ])


def build_rule_trace(df, rules_file, sheet_name, year_mode="current"):
    """v7 trace: покажува effective matching, synthetic fallback, исклучени синтетики и причина."""
    try:
        rules = read_rules(rules_file, sheet_name)
    except Exception as e:
        return pd.DataFrame([{
            "Sheet": sheet_name,
            "Грешка": str(e)
        }])

    rows = []

    for _, row in rules.iterrows():
        aop = format_aop(row.get("AOP", ""))
        pozicija = str(row.get("Pozicija", "")).strip()

        if year_mode == "previous":
            konto_text = row.get("konto_prethodna", row.get("konto", None))
            kolona = row.get("kolona_prethodna", None)
            operacija = row.get("Operacija_prethodna", row.get("Operacija", "+"))
        else:
            konto_text = row.get("konto", None)
            kolona = row.get("kolona", None)
            operacija = row.get("Operacija", "+")

        if pd.isna(konto_text) or pd.isna(kolona):
            continue

        prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
        prefixes = [x for x in prefixes if x]

        if prefixes:
            idxs, excluded_idxs, notes = _matching_effective_details(
                df,
                prefixes,
                kolona=str(kolona).strip()
            )
            match_mode = "effective"
        else:
            idxs, excluded_idxs, notes = [], [], []
            match_mode = "none"

        if idxs:
            matched_df = df.loc[idxs].copy()
            matched_konta = matched_df["konto"].astype(str).tolist()
            value = float(matched_df[str(kolona).strip()].apply(clean_number).sum())
        else:
            matched_konta = []
            value = 0.0

        if excluded_idxs:
            excluded_konta = df.loc[excluded_idxs, "konto"].astype(str).tolist()
        else:
            excluded_konta = []

        if str(operacija).strip() == "-":
            value = -value

        rows.append({
            "AOP": aop,
            "Позиција": pozicija,
            "Правило конта": str(konto_text),
            "Колона": str(kolona),
            "Операција": str(operacija),
            "Начин": match_mode,
            "Број конта": len(matched_konta),
            "Погодени конта": ", ".join(matched_konta[:40]) + (" ..." if len(matched_konta) > 40 else ""),
            "Исклучени синтетики": ", ".join(excluded_konta[:40]) + (" ..." if len(excluded_konta) > 40 else ""),
            "Објаснување": " | ".join(notes[:10]) + (" ..." if len(notes) > 10 else ""),
            "Износ": value,
            "Статус": "✅ OK" if len(matched_konta) > 0 else "⚠️ Нема погодени конта",
        })

    return pd.DataFrame(rows)

# =====================================================
# END SIGMA v6 EFFECTIVE ACCOUNT ENGINE
# =====================================================

# =====================================================
# END SIGMA v5 SMART PREFIX ENGINE FOR BS
# =====================================================

# =====================================================
# END SIGMA v4 BU PREFIX ENGINE OVERRIDES
# =====================================================

# =====================================================
# END SIGMA STABLE ENGINE OVERRIDES
# =====================================================


# =====================================================
# SIGMA v9 FIXES:
# 1) Shared synthetic fallback is consumed once per BU/BS calculation
#    (2340 + 2341 no longer duplicate parent 234 when only 234 exists).
# 2) Synthetic 019 is NOT used as fallback for 0191/0192/0193 rules;
#    when only 019 exists, it is allocated proportionally across fixed assets.
# =====================================================

def _is_analytic_019_prefix(prefix):
    prefix = normalize_rule_prefix(prefix)
    return bool(re.fullmatch(r"019\d+", prefix)) and prefix != "019"


def _fallback_key(parent, col):
    return (str(parent), str(col).strip() if col is not None else "")


def _matching_effective_indices_v9(df, prefixes, kolona=None, tolerance=1.0, used_fallback_parents=None):
    if df is None or df.empty or "konto" not in df.columns:
        return []

    prefixes = [normalize_rule_prefix(p) for p in prefixes]
    prefixes = [p for p in prefixes if p]
    if not prefixes:
        return []

    konta_norm = df["konto"].astype(str).apply(normalize_konto)
    selected = []

    col = str(kolona).strip() if kolona is not None and not pd.isna(kolona) else None
    has_col = bool(col and col in df.columns)

    if used_fallback_parents is None:
        used_fallback_parents = set()

    for prefix in prefixes:
        match_mask = konta_norm.apply(lambda k: str(k).startswith(prefix))
        match_idxs = list(df.index[match_mask])

        if not match_idxs:
            fallback_idxs, fallback_parent = _find_existing_synthetic_parent(konta_norm, prefix)

            # IMPORTANT: never map analytic 0191/0192/0193 rules to synthetic 019.
            # If only synthetic 019 exists, raspredeli_019_proporcionalno handles it once.
            if fallback_idxs and _is_analytic_019_prefix(prefix) and str(fallback_parent) == "019":
                continue

            if fallback_idxs:
                key = _fallback_key(fallback_parent, col)
                if key in used_fallback_parents:
                    continue
                used_fallback_parents.add(key)

                for fallback_idx in fallback_idxs:
                    if fallback_idx not in selected:
                        selected.append(fallback_idx)
            continue

        match_idxs = sorted(
            match_idxs,
            key=lambda idx: (len(str(konta_norm.loc[idx])), str(konta_norm.loc[idx]))
        )

        for idx in match_idxs:
            child_idxs = _immediate_child_indices(konta_norm, idx, match_idxs)

            if not child_idxs:
                if idx not in selected:
                    selected.append(idx)
                continue

            if not has_col:
                continue

            parent_value = float(df.loc[[idx], col].apply(clean_number).sum())
            children_value = float(df.loc[child_idxs, col].apply(clean_number).sum())

            parent_is_subtotal = (
                abs(parent_value - children_value) <= tolerance
                and abs(children_value) > tolerance
            )

            parent_empty_children_have_value = (
                abs(parent_value) <= tolerance
                and abs(children_value) > tolerance
            )

            if parent_is_subtotal or parent_empty_children_have_value:
                continue

            if idx not in selected:
                selected.append(idx)

    return selected


def _matching_effective_details_v9(df, prefixes, kolona=None, tolerance=1.0, used_fallback_parents=None):
    if df is None or df.empty or "konto" not in df.columns:
        return [], [], []

    prefixes = [normalize_rule_prefix(p) for p in prefixes]
    prefixes = [p for p in prefixes if p]
    if not prefixes:
        return [], [], []

    konta_norm = df["konto"].astype(str).apply(normalize_konto)
    selected, excluded, notes = [], [], []

    col = str(kolona).strip() if kolona is not None and not pd.isna(kolona) else None
    has_col = bool(col and col in df.columns)

    if used_fallback_parents is None:
        used_fallback_parents = set()

    for prefix in prefixes:
        match_mask = konta_norm.apply(lambda k: str(k).startswith(prefix))
        match_idxs = list(df.index[match_mask])

        if not match_idxs:
            fallback_idxs, fallback_parent = _find_existing_synthetic_parent(konta_norm, prefix)

            if fallback_idxs and _is_analytic_019_prefix(prefix) and str(fallback_parent) == "019":
                notes.append(
                    f"{prefix}: постои само синтетика 019 -> не се зема како директен fallback; "
                    "се распределува пропорционално еднаш"
                )
                continue

            if fallback_idxs:
                key = _fallback_key(fallback_parent, col)
                if key in used_fallback_parents:
                    notes.append(
                        f"{prefix}: синтетика {fallback_parent} веќе е земена во претходно правило -> нема дуплирање"
                    )
                    continue

                used_fallback_parents.add(key)
                for fallback_idx in fallback_idxs:
                    if fallback_idx not in selected:
                        selected.append(fallback_idx)
                notes.append(
                    f"{prefix}: нема аналитика во документот -> земена синтетика {fallback_parent}"
                )
            else:
                notes.append(f"{prefix}: нема погодени конта")
            continue

        match_idxs = sorted(
            match_idxs,
            key=lambda idx: (len(str(konta_norm.loc[idx])), str(konta_norm.loc[idx]))
        )

        for idx in match_idxs:
            konto = str(konta_norm.loc[idx])
            child_idxs = _immediate_child_indices(konta_norm, idx, match_idxs)

            if not child_idxs:
                if idx not in selected:
                    selected.append(idx)
                continue

            if not has_col:
                if idx not in excluded:
                    excluded.append(idx)
                notes.append(f"{konto}: има аналитика -> синтетиката е исклучена")
                continue

            parent_value = float(df.loc[[idx], col].apply(clean_number).sum())
            children_value = float(df.loc[child_idxs, col].apply(clean_number).sum())

            parent_is_subtotal = (
                abs(parent_value - children_value) <= tolerance
                and abs(children_value) > tolerance
            )
            parent_empty_children_have_value = (
                abs(parent_value) <= tolerance
                and abs(children_value) > tolerance
            )

            if parent_is_subtotal:
                if idx not in excluded:
                    excluded.append(idx)
                notes.append(f"{konto}: синтетика=subtotal -> земени аналитики")
                continue

            if parent_empty_children_have_value:
                if idx not in excluded:
                    excluded.append(idx)
                notes.append(f"{konto}: синтетика=0 -> земени аналитики")
                continue

            if idx not in selected:
                selected.append(idx)
            notes.append(f"{konto}: директно книжење + аналитики -> земени и двете")

    return selected, excluded, notes


def sum_rule_effective_context(df, konto_text, kolona, operacija, used_fallback_parents=None):
    if pd.isna(konto_text) or pd.isna(kolona):
        return 0.0

    kolona = str(kolona).strip()
    if kolona not in df.columns:
        return 0.0

    prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
    prefixes = [x for x in prefixes if x]
    if not prefixes:
        return 0.0

    idxs = _matching_effective_indices_v9(
        df,
        prefixes,
        kolona=kolona,
        used_fallback_parents=used_fallback_parents
    )
    if not idxs:
        return 0.0

    value = float(df.loc[idxs, kolona].apply(clean_number).sum())
    return -value if str(operacija).strip() == "-" else value


# Contextless fallback for modules that call the old name directly.
def sum_rule_effective(df, konto_text, kolona, operacija):
    return sum_rule_effective_context(df, konto_text, kolona, operacija, used_fallback_parents=None)


def sum_rule(df, konto_text, kolona, operacija):
    return sum_rule_effective(df, konto_text, kolona, operacija)


def sum_rule_prefix_all(df, konto_text, kolona, operacija):
    return sum_rule_effective(df, konto_text, kolona, operacija)


def sum_rule_smart_prefix(df, konto_text, kolona, operacija):
    return sum_rule_effective(df, konto_text, kolona, operacija)


def raspredeli_019_proporcionalno(df, values, kolona_019, kolona_bruto):
    """
    v9: Ако има само синтетика 019, не ја доделуваме целата на едно AOP.
    Ја распределуваме пропорционално по бруто вредност на средствата.
    Ако постојат аналитики 0191/0192/0193..., НЕ правиме распределба,
    бидејќи правилата веќе точно ја носат амортизацијата.
    """
    if not ima_zbirno_019(df):
        return values

    if ima_analitika_019(df):
        return values

    iznos_019 = abs(suma_konto_019_exact(df, "019", kolona_019))
    if iznos_019 == 0:
        return values

    # AOP групи што може да носат амортизација кога имаме само збирно 019.
    # Земјиште/аванси/подготовка не ги вклучуваме во базата.
    depreciable_aops = ["004", "012", "013", "014", "015", "016", "019"]

    bruto_values = {}
    vkupno_bruto = 0.0

    for aop in depreciable_aops:
        bruto = float(values.get(aop, 0.0))
        if bruto > 0:
            bruto_values[aop] = bruto
            vkupno_bruto += bruto

    if vkupno_bruto == 0:
        return values

    for aop, bruto in bruto_values.items():
        del_od_019 = iznos_019 * (bruto / vkupno_bruto)
        values[aop] = values.get(aop, 0.0) - del_od_019

    return values


def calculate_bilans_uspeh(df, rules_file):
    """v9: BU со context-aware fallback: 234 не се дуплира за 2340 и 2341."""
    rules = read_rules(rules_file, "Pravila Bilans Uspeh")
    values, positions, order = {}, {}, []
    used_fallback_parents = set()

    for _, row in rules.iterrows():
        aop = format_aop(row["AOP"])
        if aop not in order:
            order.append(aop)
        positions[aop] = row.get("Pozicija", "")

        val = sum_rule_effective_context(
            df,
            row.get("konto", None),
            row.get("kolona", None),
            row.get("Operacija", "+"),
            used_fallback_parents=used_fallback_parents
        )
        values[aop] = values.get(aop, 0.0) + val

    values = apply_formulas_and_logic(rules, values)

    return pd.DataFrame([
        {"AOP": aop, "Pozicija": positions.get(aop, ""), "Iznos": values.get(aop, 0.0)}
        for aop in order
    ])


def calculate_bilans_sostojba(df, rules_file, bu=None):
    """v9: BS со context-aware fallback + исправена 019 пропорционална распределба."""
    rules = read_rules(rules_file, "Pravila bilans Sostojba")
    current_values, previous_values, positions, order = {}, {}, {}, []
    current_used_fallbacks = set()
    previous_used_fallbacks = set()

    bu_current_map = {}
    if bu is not None:
        for _, bu_row in bu.iterrows():
            bu_aop = str(bu_row["AOP"]).strip().replace(".0", "").zfill(3)
            bu_current_map[f"BU{bu_aop}"] = float(bu_row["Iznos"])

    for _, row in rules.iterrows():
        aop = format_aop(row["AOP"])
        if aop not in order:
            order.append(aop)
        positions[aop] = row.get("Pozicija", "")

        current_val = sum_rule_effective_context(
            df,
            row.get("konto", None),
            row.get("kolona", None),
            row.get("Operacija", "+"),
            used_fallback_parents=current_used_fallbacks
        )

        previous_val = sum_rule_effective_context(
            df,
            row.get("konto_prethodna", row.get("konto", None)),
            row.get("kolona_prethodna", None),
            row.get("Operacija_prethodna", row.get("Operacija", "+")),
            used_fallback_parents=previous_used_fallbacks
        )

        current_values[aop] = current_values.get(aop, 0.0) + current_val
        previous_values[aop] = previous_values.get(aop, 0.0) + previous_val

    if bu is not None:
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)

    current_values = raspredeli_019_proporcionalno(df, current_values, "krajno_pobaruva", "krajno_dolzi")
    current_values = apply_formulas_and_logic(rules, current_values, external_values=bu_current_map)

    if bu is not None:
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)
        current_values = apply_formulas_and_logic(rules, current_values, external_values=bu_current_map)
        current_values["077"] = bu_current_map.get("BU255", 0.0)
        current_values["078"] = bu_current_map.get("BU256", 0.0)

    previous_values = raspredeli_019_proporcionalno(df, previous_values, "prethodna_pobaruva", "prethodna_dolzi")
    previous_values = apply_formulas_and_logic(rules, previous_values)

    return pd.DataFrame([
        {
            "AOP": aop,
            "Pozicija": positions.get(aop, ""),
            "Tekovna godina": current_values.get(aop, 0.0),
            "Prethodna godina": previous_values.get(aop, 0.0)
        }
        for aop in order
    ])


def build_rule_trace(df, rules_file, sheet_name, year_mode="current"):
    """v9 trace: покажува кога fallback синтетика е искористена и кога е спречено дуплирање."""
    try:
        rules = read_rules(rules_file, sheet_name)
    except Exception as e:
        return pd.DataFrame([{"Sheet": sheet_name, "Грешка": str(e)}])

    rows = []
    used_fallback_parents = set()

    for _, row in rules.iterrows():
        aop = format_aop(row.get("AOP", ""))
        pozicija = str(row.get("Pozicija", "")).strip()

        if year_mode == "previous":
            konto_text = row.get("konto_prethodna", row.get("konto", None))
            kolona = row.get("kolona_prethodna", None)
            operacija = row.get("Operacija_prethodna", row.get("Operacija", "+"))
        else:
            konto_text = row.get("konto", None)
            kolona = row.get("kolona", None)
            operacija = row.get("Operacija", "+")

        if pd.isna(konto_text) or pd.isna(kolona):
            continue

        prefixes = [normalize_rule_prefix(x) for x in konto_list(konto_text)]
        prefixes = [x for x in prefixes if x]

        if prefixes:
            idxs, excluded_idxs, notes = _matching_effective_details_v9(
                df,
                prefixes,
                kolona=str(kolona).strip(),
                used_fallback_parents=used_fallback_parents
            )
            match_mode = "effective-v9"
        else:
            idxs, excluded_idxs, notes = [], [], []
            match_mode = "none"

        if idxs:
            matched_df = df.loc[idxs].copy()
            matched_konta = matched_df["konto"].astype(str).tolist()
            value = float(matched_df[str(kolona).strip()].apply(clean_number).sum())
        else:
            matched_konta = []
            value = 0.0

        excluded_konta = df.loc[excluded_idxs, "konto"].astype(str).tolist() if excluded_idxs else []

        if str(operacija).strip() == "-":
            value = -value

        rows.append({
            "AOP": aop,
            "Позиција": pozicija,
            "Правило конта": str(konto_text),
            "Колона": str(kolona),
            "Операција": str(operacija),
            "Начин": match_mode,
            "Број конта": len(matched_konta),
            "Погодени конта": ", ".join(matched_konta[:40]) + (" ..." if len(matched_konta) > 40 else ""),
            "Исклучени синтетики": ", ".join(excluded_konta[:40]) + (" ..." if len(excluded_konta) > 40 else ""),
            "Објаснување": " | ".join(notes[:10]) + (" ..." if len(notes) > 10 else ""),
            "Износ": value,
            "Статус": "✅ OK" if len(matched_konta) > 0 else "⚠️ Нема погодени конта",
        })

    return pd.DataFrame(rows)

# =====================================================
# END SIGMA v9 FIXES
# =====================================================

@st.cache_data(show_spinner=False, max_entries=3)
def load_zaklucen_cached(file_name, file_bytes):
    """
    Го чита заклучниот лист само еднаш за ист uploaded фајл.
    При кликање низ sidebar Streamlit rerun веќе не го парсира PDF/Excel повторно.
    """
    file_obj = BytesIO(file_bytes)
    file_obj.name = file_name

    if file_name.lower().endswith(".pdf"):
        df_loaded = read_zaklucen_pdf(file_obj)
    else:
        df_loaded = read_zaklucen(file_obj)

    return normalize_dataframe_numbers(df_loaded)



# =====================================================
# CFO REPORT - HTML / PRINT TO PDF
# Додадено врз стабилниот работен код без промена на BU/BS/019/234/240 логиката.
# =====================================================

def cfo_escape(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return html_lib.escape(str(value))


def cfo_format_number(value):
    try:
        if pd.isna(value):
            return ""
        return f"{float(value):,.0f}".replace(",", ".")
    except Exception:
        return cfo_escape(value)


def cfo_format_percent(value):
    try:
        if pd.isna(value):
            return ""
        return f"{float(value) * 100:.2f}%"
    except Exception:
        return cfo_escape(value)


def cfo_remove_empty_report_rows(df, value_columns):
    """
    Ги трга нултите/празните редови само за приказ во CFO извештај.
    Не ја менува пресметковната логика на BU, BS, 019, 234/240.
    """
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()

    result = df.copy()
    existing_value_cols = [c for c in value_columns if c in result.columns]

    if not existing_value_cols:
        return result

    numeric_part = result[existing_value_cols].apply(
        pd.to_numeric,
        errors="coerce"
    ).fillna(0)

    mask_has_value = numeric_part.abs().sum(axis=1) != 0
    return result.loc[mask_has_value].copy()


def cfo_prepare_df_for_html(df, columns=None):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()

    display_df = df.copy()

    if columns:
        existing_cols = [c for c in columns if c in display_df.columns]
        if existing_cols:
            display_df = display_df[existing_cols]

    for col in display_df.columns:
        if pd.api.types.is_numeric_dtype(display_df[col]):
            display_df[col] = display_df[col].apply(cfo_format_number)
        else:
            display_df[col] = display_df[col].apply(cfo_escape)

    return display_df


def cfo_df_to_html(df, title, columns=None):
    display_df = cfo_prepare_df_for_html(df, columns)

    if display_df.empty:
        return f"""
        <section class="report-section page-break">
            <h2>{cfo_escape(title)}</h2>
            <p class="muted">Нема достапни податоци за овој извештај.</p>
        </section>
        """

    table_html = display_df.to_html(
        index=False,
        border=0,
        escape=False,
        classes="cfo-table"
    )

    return f"""
    <section class="report-section page-break">
        <h2>{cfo_escape(title)}</h2>
        {table_html}
    </section>
    """


def cfo_get_aop_value(report_df, aop, value_col="Iznos"):
    try:
        if report_df is None or report_df.empty:
            return 0.0
        if "AOP" not in report_df.columns or value_col not in report_df.columns:
            return 0.0

        aop_series = (
            report_df["AOP"]
            .astype(str)
            .str.strip()
            .str.replace(".0", "", regex=False)
            .str.zfill(3)
        )

        return float(
            pd.to_numeric(
                report_df.loc[aop_series == str(aop).zfill(3), value_col],
                errors="coerce"
            ).fillna(0).sum()
        )
    except Exception:
        return 0.0


def cfo_get_kpi_raw(kpi_df, naziv_contains):
    try:
        if kpi_df is None or kpi_df.empty:
            return 0.0

        if "Naziv" not in kpi_df.columns:
            return 0.0

        match = kpi_df[
            kpi_df["Naziv"].astype(str).str.contains(
                naziv_contains,
                case=False,
                na=False
            )
        ]

        if match.empty:
            return 0.0

        if "RawValue" in match.columns:
            return float(match.iloc[0].get("RawValue", 0.0))

        return 0.0
    except Exception:
        return 0.0


def build_cfo_snapshot_df_for_report(df, bu, kpi):
    try:
        be_cfo = calculate_break_even(
            df=df,
            bu=bu,
            sales_aops=["202"],
            pct_402=70,
            pct_403=70
        )
        be_map = dict(zip(be_cfo["Показател"], be_cfo["Вредност"]))
    except Exception:
        be_map = {}

    sales_revenue = float(be_map.get("Приход од продажба", 0))
    break_even_sales = float(be_map.get("Break-even продажба", 0))
    margin_of_safety = float(be_map.get("Над Break-even", 0))
    safety_margin_pct = float(be_map.get("Безбедносна маргина %", 0))

    net_profit = cfo_get_aop_value(bu, "255", "Iznos")
    net_loss = cfo_get_aop_value(bu, "256", "Iznos")
    net_result = net_profit - net_loss

    current_ratio = cfo_get_kpi_raw(kpi, "ТЕКОВНА ЛИКВИДНОСТ")
    profit_margin = cfo_get_kpi_raw(kpi, "ПРОФИТНА МАРЖА")
    debt_ratio = cfo_get_kpi_raw(kpi, "ВКУПНА ЗАДОЛЖЕНОСТ")

    score = 0

    if net_result > 0 and profit_margin >= 0.15:
        score += 35
    elif net_result > 0 and profit_margin >= 0.05:
        score += 25
    elif net_result > 0:
        score += 15

    if current_ratio >= 1.5:
        score += 35
    elif current_ratio >= 1.0:
        score += 25
    elif current_ratio >= 0.7:
        score += 15
    else:
        score += 5

    if safety_margin_pct >= 0.30:
        score += 30
    elif safety_margin_pct >= 0.10:
        score += 20
    elif safety_margin_pct > 0:
        score += 10

    if score >= 80:
        status = "🟢 Стабилна финансиска состојба"
        comment = "Компанијата има добра комбинација на профитабилност, ликвидност и безбедносна маргина."
    elif score >= 55:
        status = "🟡 Средна зона"
        comment = "Компанијата е функционална, но има области каде што треба да се следи ликвидноста, профитабилноста или break-even позицијата."
    else:
        status = "🔴 Потребно е внимание"
        comment = "Потребна е подетална CFO анализа, особено на трошоци, ликвидност и продажба."

    return pd.DataFrame([
        {"Показател": "CFO Score", "Вредност": f"{score}/100", "Коментар": status},
        {"Показател": "Нето резултат", "Вредност": cfo_format_number(net_result), "Коментар": "Добивка - загуба"},
        {"Показател": "Приход од продажба", "Вредност": cfo_format_number(sales_revenue), "Коментар": "Основен приход за анализа"},
        {"Показател": "Break-even продажба", "Вредност": cfo_format_number(break_even_sales), "Коментар": "Минимална продажба за покривање на трошоци"},
        {"Показател": "Над Break-even", "Вредност": cfo_format_number(margin_of_safety), "Коментар": "Разлика меѓу продажба и break-even"},
        {"Показател": "Безбедносна маргина", "Вредност": cfo_format_percent(safety_margin_pct), "Коментар": "Колку компанијата е над критичната точка"},
        {"Показател": "Тековна ликвидност", "Вредност": f"{current_ratio:.2f}", "Коментар": "Способност за покривање краткорочни обврски"},
        {"Показател": "Профитна маржа", "Вредност": cfo_format_percent(profit_margin), "Коментар": "Профитабилност од продажбата"},
        {"Показател": "Вкупна задолженост", "Вредност": cfo_format_percent(debt_ratio), "Коментар": "Однос на обврски во структурата"},
        {"Показател": "CFO Коментар", "Вредност": status, "Коментар": comment},
    ])


def build_cfo_report_html(company_name, period_label, sections_html):
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    all_sections = "\n".join(sections_html)

    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>CFO Извештај</title>
<style>
    body {{
        font-family: Arial, sans-serif;
        margin: 36px;
        color: #111827;
        background: white;
    }}

    .print-button {{
        position: fixed;
        top: 16px;
        right: 16px;
        background: #111827;
        color: white;
        border: none;
        border-radius: 8px;
        padding: 10px 16px;
        font-weight: 700;
        cursor: pointer;
        z-index: 999;
    }}

    .cover {{
        min-height: 760px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        border: 2px solid #111827;
        padding: 50px;
        margin-bottom: 40px;
    }}

    .brand {{
        font-size: 18px;
        font-weight: 700;
        letter-spacing: 1px;
        color: #2563eb;
        margin-bottom: 20px;
    }}

    h1 {{
        font-size: 42px;
        margin: 0 0 12px 0;
        color: #111827;
    }}

    .subtitle {{
        font-size: 18px;
        color: #4b5563;
        line-height: 1.6;
    }}

    .meta {{
        margin-top: 70px;
        font-size: 14px;
        color: #374151;
    }}

    .report-section {{
        margin-top: 28px;
        margin-bottom: 36px;
    }}

    h2 {{
        font-size: 22px;
        color: #111827;
        border-bottom: 2px solid #111827;
        padding-bottom: 8px;
        margin-bottom: 14px;
    }}

    .cfo-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 11px;
        margin-top: 10px;
    }}

    .cfo-table th {{
        background: #f3f4f6;
        color: #111827;
        border: 1px solid #d1d5db;
        padding: 7px;
        text-align: left;
        font-weight: 700;
    }}

    .cfo-table td {{
        border: 1px solid #e5e7eb;
        padding: 6px;
        vertical-align: top;
    }}

    .cfo-table tr:nth-child(even) td {{
        background: #fafafa;
    }}

    .muted {{
        color: #6b7280;
        font-size: 13px;
    }}

    .footer {{
        margin-top: 40px;
        padding-top: 12px;
        border-top: 1px solid #e5e7eb;
        font-size: 11px;
        color: #6b7280;
        text-align: center;
    }}

    @media print {{
        .print-button {{
            display: none;
        }}

        body {{
            margin: 18px;
        }}

        .page-break {{
            page-break-before: always;
        }}

        .cover {{
            page-break-after: always;
            min-height: 700px;
        }}
    }}
</style>
</head>
<body>

<button class="print-button" onclick="window.print()">🖨️ Print / Save as PDF</button>

<div class="cover">
    <div class="brand">SIGMA ACCOUNTING</div>
    <h1>📘 CFO Извештај</h1>
    <div class="subtitle">
        <strong>{cfo_escape(company_name)}</strong><br>
        Период: {cfo_escape(period_label)}
    </div>

    <div class="meta">
        Извештајот е подготвен врз основа на прикачен заклучен лист и автоматски генерирани финансиски извештаи.<br>
        Генерирано на: {generated_at}
    </div>
</div>

{all_sections}

<div class="footer">
    CFO Извештај генериран преку SIGMA Financial Reports.
</div>

</body>
</html>
"""


if zaklucen_file :
    try:
        # FINALNA SIGURNOST + SPEED:
        # PDF/Excel се чита само еднаш за истиот upload фајл.
        df = load_zaklucen_cached(zaklucen_file.name, zaklucen_file.getvalue())

        # Подготовка за приказ/export. Се пресметува еднаш, но се прикажува
        # само кога корисникот е на „Заклучен лист“.
        sum_row = {"konto": "", "naziv": "VKUPNO"}
        for col in df.columns:
            if col not in ["konto", "naziv"]:
                sum_row[col] = int(df[col].sum())
        df_total = pd.concat([df, pd.DataFrame([sum_row])], ignore_index=True)
        for col in df_total.columns:
            if col not in ["konto", "naziv"]:
                df_total.loc[:, col] = df_total[col].round(0).astype(int)

        def style_total_row(row):
            if row["naziv"] == "VKUPNO":
                return ["font-weight: bold; font-size: 18px; background-color: #d9d9d9;" for _ in row]
            return ["" for _ in row]

        if izbran_izvestaj == "🏠 Заклучен лист":
            st.subheader("📑 Заклучен лист")
            st.dataframe(
                df_total.style.apply(style_total_row, axis=1),
                use_container_width=True
            )

            st.subheader("✅ Контрола на заклучен лист")
            zaklucen_kontrola = validate_zaklucen_list(df)
            if not zaklucen_kontrola.empty:
                st.dataframe(zaklucen_kontrola, use_container_width=True)
                open_diffs = zaklucen_kontrola[zaklucen_kontrola["Razlika"] != 0]
                if open_diffs.empty:
                    st.success("Заклучниот лист е затворен ✅ Должи = Побарува")
                else:
                    for _, kontrola_row in open_diffs.iterrows():
                        st.error(
                            f'{kontrola_row["Kontrola"]}: разлика {kontrola_row["Razlika"]:,.0f}'
                        )

            st.subheader("🧮 Збир по класи од заклучниот лист")
            class_summary = build_konto_class_summary(df, mode="raw")
            class_summary_display = class_summary.copy()
            for _col in class_summary_display.columns:
                if _col not in ["Класа", "Назив", "Број на редови"]:
                    class_summary_display.loc[:, _col] = class_summary_display[_col].round(0).astype(int)
            st.dataframe(class_summary_display, use_container_width=True)

            with st.expander("🔎 Збир по класи за правила - без дуплирање синтетика/аналитика", expanded=False):
                st.caption(
                    "Оваа табела служи за проверка на правилата. Ако постои синтетичко конто и негови аналитики, "
                    "се зема синтетиката за да не се дуплираат износите."
                )
                class_summary_rules = build_konto_class_summary(df, mode="rules")
                class_summary_rules_display = class_summary_rules.copy()
                for _col in class_summary_rules_display.columns:
                    if _col not in ["Класа", "Назив", "Број на редови"]:
                        class_summary_rules_display.loc[:, _col] = class_summary_rules_display[_col].round(0).astype(int)
                st.dataframe(class_summary_rules_display, use_container_width=True)

        bu = calculate_bilans_uspeh(df, pravila_file)
        bs = calculate_bilans_sostojba(df, pravila_file, bu)

        if izbran_izvestaj == "🏠 Заклучен лист" and ima_zbirno_019(df) and not ima_analitika_019(df):
            st.warning(
                "⚠️ Детектирано е збирно конто 019 без аналитика 0191/0192/0193. "
                "Исправката на вредност 019 е распределена пропорционално по материјални средства со бруто вредност > 0."
            )

        bu255 = float(bu.loc[bu["AOP"].astype(str).str.zfill(3) == "255", "Iznos"].sum())
        bu256 = float(bu.loc[bu["AOP"].astype(str).str.zfill(3) == "256", "Iznos"].sum())
        #st.info(
            #f"BU rezultatot e direktno prenesen vo BS: "
            #f"BU255 → BS077 = {bu255:,.0f}; BU256 → BS078 = {bu256:,.0f}. "
            #f"Nema avtomatsko zatvoranje so razlika — kontrolata Aktiva = Pasiva ostanuva realna."
        #)

      
        sd = calculate_seopfatna_dobivka(bu, pravila_file)
        pravila_val = pd.read_excel(
        pravila_file,
    sheet_name="Pravila Valuation"
)

# ги чисти празните места од имињата на колоните
        pravila_val.columns = pravila_val.columns.str.strip()

        valuation_rules = {}

        for _, row in pravila_val.iterrows():
            valuation_rules[str(row["Pozicija"]).strip()] = row["AOP"]

        #st.write("Valuation rules:", valuation_rules)

        # ===== VALUES DICT =====
        values_dict = {}

        # BS values
        for _, row in bs.iterrows():
            aop = str(row["AOP"]).strip()
            values_dict[aop + "_T"] = row["Tekovna godina"]

        # SD values
        for _, row in sd.iterrows():
            red = str(row["Red"]).strip()

            values_dict[red] = row["Tekovna godina"]

            # dodatna sigurnost ako nekade e samo broj
            if red.isdigit():
                values_dict["SD" + red] = row["Tekovna godina"]

        #st.write("VALUES DICT KEYS:", list(values_dict.keys()))
        #st.write("SD14 value:", values_dict.get("SD14"))
        #st.write("SD8 value:", values_dict.get("SD8"))

        # ===== VALUATION TEST =====
        ebitda = calculate_formula(
            valuation_rules["EBITDA"],
            values_dict
        )

        debt = calculate_formula(
            valuation_rules["Debt"],
            values_dict
        )

        cash = calculate_formula(
            valuation_rules["Cash"],
            values_dict
        )

        #st.write("EBITDA:", ebitda)
        #st.write("Debt:", debt)
        #st.write("Cash:", cash)

        try:
            cf = calculate_cash_flow(df, pravila_file, bu, bs)
        except Exception as cash_error:
            cf = None
            st.warning(f"Cash Flow ne e vcitan: {cash_error}")

        with st.sidebar.expander("⚙️ KPI параметри", expanded=False):
            with st.form("kpi_params_form", clear_on_submit=False):
                broj_vraboteni = st.number_input(
                    "Број на вработени",
                    min_value=0,
                    value=0,
                    step=1,
                    key="broj_vraboteni_kpi"
                )

                meseci_rabotenje = st.number_input(
                    "Месеци на работење",
                    min_value=0,
                    max_value=12,
                    value=12,
                    step=1,
                    key="meseci_rabotenje_kpi"
                )

                st.form_submit_button(
                    "✅ Примени KPI параметри",
                    use_container_width=True
                )
            st.caption("Промените се пресметуваат само по клик на Примени.")

        # Не го менуваме оригиналниот BU што го користат останатите извештаи.
        # KPI добива посебна копија со оперативните параметри.
        bu_kpi = bu.copy()
        bu_kpi.loc[bu_kpi["AOP"] == "257", "Iznos"] = broj_vraboteni
        bu_kpi.loc[bu_kpi["AOP"] == "258", "Iznos"] = meseci_rabotenje

        try:
            kpi = calculate_kpi(
                pravila_file,
                bu_kpi,
                bs,
                broj_vraboteni=broj_vraboteni,
                meseci_rabotenje=meseci_rabotenje
            )
        except Exception as kpi_error:
            kpi = None
            st.warning(f"KPI ne e vcitan: {kpi_error}")

        bu["Iznos"] = bu["Iznos"].round(0).astype(int)
        bs["Tekovna godina"] = bs["Tekovna godina"].round(0).astype(int)
        bs["Prethodna godina"] = bs["Prethodna godina"].round(0).astype(int)
        if cf is not None:
            cf["Iznos"] = cf["Iznos"].round(0).astype(int)

        aktiva = bs.loc[bs["AOP"] == "063", "Tekovna godina"].sum()
        pasiva = bs.loc[bs["AOP"] == "111", "Tekovna godina"].sum()
        razlika = aktiva - pasiva
        aktiva_prev = bs.loc[bs["AOP"] == "063", "Prethodna godina"].sum()
        pasiva_prev = bs.loc[bs["AOP"] == "111", "Prethodna godina"].sum()
        razlika_prev = aktiva_prev - pasiva_prev

        if izbran_izvestaj == "🏠 Заклучен лист":
            st.subheader("✅ Контролни проверки на Биланс на состојба")
            if razlika == 0:
                st.success(f"Tekovna godina: Aktiva = Pasiva ✅ ({aktiva:,.0f})")
            else:
                st.error(f"Tekovna godina: Razlika ❌ {razlika:,.0f}")
            if razlika_prev == 0:
                st.success(f"Prethodna godina: Aktiva = Pasiva ✅ ({aktiva_prev:,.0f})")
            else:
                st.error(f"Prethodna godina: Razlika ❌ {razlika_prev:,.0f}")

            st.subheader("🔎 Дијагностика: класи ↔ BU/BS")
            bs_diagnostic = build_bs_class_diagnostic(df, bu, bs)
            bs_diagnostic_display = bs_diagnostic.copy()
            for _col in ["Износ 1", "Износ 2", "Разлика"]:
                if _col in bs_diagnostic_display.columns:
                    bs_diagnostic_display[_col] = bs_diagnostic_display[_col].round(0).astype(int)
            st.dataframe(bs_diagnostic_display, use_container_width=True)

            with st.expander("🔎 Дијагностика на трансфер BU/BS - кои конта влегуваат во секој AOP", expanded=False):
                st.caption(
                    "Ова е техничка трага на правилата. Ако AOP е 0, тука веднаш се гледа дали правилото не пронашло конта, "
                    "дали е погрешна колоната или дали постои дуплирање синтетика/аналитика."
                )
                bu_rule_trace = build_rule_trace(df, pravila_file, "Pravila Bilans Uspeh")
                st.markdown("**БУ - правила за трансфер**")
                if not bu_rule_trace.empty and "Износ" in bu_rule_trace.columns:
                    bu_rule_trace_display = bu_rule_trace.copy()
                    bu_rule_trace_display["Износ"] = bu_rule_trace_display["Износ"].round(0).astype(int)
                    st.dataframe(bu_rule_trace_display, use_container_width=True)
                else:
                    st.dataframe(bu_rule_trace, use_container_width=True)

                bs_rule_trace = build_rule_trace(df, pravila_file, "Pravila bilans Sostojba", year_mode="current")
                st.markdown("**БС - правила за трансфер / тековна година**")
                if not bs_rule_trace.empty and "Износ" in bs_rule_trace.columns:
                    bs_rule_trace_display = bs_rule_trace.copy()
                    bs_rule_trace_display["Износ"] = bs_rule_trace_display["Износ"].round(0).astype(int)
                    st.dataframe(bs_rule_trace_display, use_container_width=True)
                else:
                    st.dataframe(bs_rule_trace, use_container_width=True)

        def premium_lock(module_name):
            #st.warning(f"Modulot '{module_name} е достапен само во PREMIUM верзија. За пристап, ве молиме контактирајте не на [email protected]")
            st.info(" Овој модул содржи напредни функции и анализи кои се достапни само за корисниците со PREMIUM пристап. За повеќе информации за тоа како да добиете пристап до PREMIUM верзијата, контактирајте не на [email sigmaakaunting@gmail.com]")
            st.button("Upgrade to PREMIUM - {module_name}",
                      disabled=True,
                      key=f"upgrade_{module_name}"
                      )
            
        def blur_numbers_df(df):
            demo = df.copy()

            for col in demo.columns:
                if col not in ["AOP", "Pozicija", "Naziv", "Red", "Kategorija", "Opis", "Status"]:
                    demo[col] = "••••••"

            return demo    
            
        
        st.divider()

        if izbran_izvestaj == "🏠 Заклучен лист":
            pass

        elif izbran_izvestaj == "📈 Биланс на успех":
            st.subheader("Bilans na uspeh po AOP")
            st.dataframe(bu.style.apply(style_formula_rows, axis=1), use_container_width=True)

        elif izbran_izvestaj == "🏦 Биланс на состојба":
            st.subheader("Bilans na sostojba po AOP")
            st.dataframe(bs.style.apply(style_formula_rows, axis=1), use_container_width=True)

        elif izbran_izvestaj == "💧 Cash Flow":
            st.subheader("Cash Flow")

            if cf is not None:

                if user_plan == "PREMIUM":
                    st.dataframe(
                        cf.style.apply(style_cashflow_rows, axis=1),
                        use_container_width=True
                    )

                else:
                    st.warning("🔒 Cash Flow е PREMIUM модул. Бројките се скриени.")

                    cf_demo = cf.copy()
                    cf_demo["Iznos"] = "••••••"

                    st.dataframe(
                        cf_demo,
                        use_container_width=True
                    )

                    premium_lock("Cash Flow")

            else:
                st.warning("Cash Flow ne e vcitan.")

        elif izbran_izvestaj == "📊 KPI Dashboard":
            st.subheader("KPI Dashboard")

            if kpi is not None:

                for category in kpi["Kategorija"].unique():

                    st.subheader(category)

                    category_df = kpi[
                        kpi["Kategorija"] == category
                    ].reset_index(drop=True)

                    cols = st.columns(4)

                    for i in range(len(category_df)):

                        row = category_df.iloc[i]

                        with cols[i % 4]:

                            st.metric(
                                label=str(row["Naziv"]),
                                value=str(row["Vrednost"])
                            )

                st.subheader("🧠 Автоматски коментари")

                for _, row in kpi.iterrows():

                    status = str(row.get("Status", ""))
                    naziv = str(row.get("Naziv", ""))
                    vrednost = str(row.get("Vrednost", ""))

                    if "🟢" in status:
                        st.success(f"{status} {naziv}: {vrednost} — показателот е во добра зона.")

                    elif "🔴" in status:
                        st.error(f"{status} {naziv}: {vrednost} — потребно е внимание и дополнителна анализа.")

                    elif "🟡" in status:
                        st.warning(f"{status} {naziv}: {vrednost} — показателот е во средна зона.")

                    else:
                        st.info(f"{naziv}: {vrednost}")

                with st.expander("📋 Детално објаснување на KPI"):
                    st.dataframe(kpi, use_container_width=True)

            else:
                st.warning("KPI ne e vcitan.")

        elif izbran_izvestaj == "🧾 Сеопфатна добивка":
            st.subheader("📊 Извештај за сеопфатна добивка")

            display_sd = sd[["Red", "Naziv", "Tekovna godina"]]

            def highlight_totals(row):
                if str(sd.loc[row.name, "Stil"]).lower() == "total":
                    return [
                        "font-weight: bold; font-size: 18px;"
                        for _ in row
                    ]

                return ["" for _ in row]

            styled_sd = (
                display_sd.style
                .format({
                    "Tekovna godina": "{:,.2f}"
                })
                .apply(highlight_totals, axis=1)
            )

            st.dataframe(
                styled_sd,
                use_container_width=True,
                height=700
            )

        elif izbran_izvestaj == "📍 Break-even Analysis":
            st.subheader("📍 Break-even Analysis")

            with st.sidebar.expander("⚙️ Break-even параметри", expanded=False):
                with st.form("break_even_params_form", clear_on_submit=False):
                    sales_aop_input = st.text_input(
                        "AOP за приходи од продажба",
                        value="202",
                        key="be_sales_aop"
                    )

                    months_in_period = st.number_input(
                        "Број на месеци во извештајот",
                        min_value=1,
                        max_value=12,
                        value=12,
                        step=1,
                        key="be_months"
                    )

                    pct_402 = st.slider(
                        "Варијабилен дел од 402 - струја (%)",
                        0,
                        100,
                        70,
                        key="be_pct_402"
                    )

                    pct_403 = st.slider(
                        "Варијабилен дел од 403 - енергија (%)",
                        0,
                        100,
                        70,
                        key="be_pct_403"
                    )

                    include_701 = st.checkbox(
                        "Вклучи 701 - набавна вредност на продадени стоки",
                        value=True,
                        key="be_include_701",
                        help="За трговски фирми 701 најчесто е варијабилен трошок и треба да влезе во break-even пресметката."
                    )

                    # Slider-от е секогаш видлив; ако 701 е исклучено, вредноста
                    # едноставно не се користи. Така формата не бара дополнителен rerun.
                    pct_701_input = st.slider(
                        "Варијабилен дел од 701 - набавна вредност (%)",
                        0,
                        100,
                        100,
                        key="be_pct_701"
                    )

                    st.form_submit_button(
                        "✅ Примени Break-even параметри",
                        use_container_width=True
                    )

                st.caption("Промените се пресметуваат само по клик на Примени.")

            pct_701 = pct_701_input if include_701 else 0

            sales_aops = [
                x.strip()
                for x in sales_aop_input.split(",")
                if x.strip() != ""
            ]

            be = calculate_break_even(
                df=df,
                bu=bu,
                sales_aops=sales_aops,
                pct_402=pct_402,
                pct_403=pct_403,
                months_in_period=months_in_period,
                include_701=include_701,
                pct_701=pct_701
            )
            
            value_map = dict(zip(be["Показател"], be["Вредност"]))

            c1, c2, c3, c4, c5 = st.columns(5)

            c1.metric(
                "Приход од продажба",
                f'{value_map.get("Приход од продажба", 0):,.0f}'
            )

            c2.metric(
                "Break-even продажба",
                f'{value_map.get("Break-even продажба", 0):,.0f}'
            )

            c3.metric(
                "Над Break-even ",
                f'{value_map.get("Над Break-even", 0):,.0f}'
            )

            c4.metric(
                "Безбедносна маргина",
                f'{value_map.get("Безбедносна маргина %", 0) * 100:.2f}%'
            )

            c5.metric(
                "Месечен Break-even",
                f'{value_map.get("Просечен месечен Break-even", 0):,.0f}'
            )

            st.subheader("📋 Детална пресметка")

            #st.write(be)
            #st.write(contribution_margin_ratio)
            display_be = be.copy()

            for i, row in display_be.iterrows():
                pokazatel = str(row["Показател"]).strip()

                if pokazatel in ["Варијабилни трошоци %", "Маржа за покривање на фиксни трошоци", "Безбедносна маргина %"]:
                    display_be.loc[i, "Вредност"] = f'{float(row["Вредност"]) * 100:.2f}%'
                elif pokazatel not in ["Статус", "Коментар"]:
                    display_be.loc[i, "Вредност"] = f'{float(row["Вредност"]):,.0f}'
            #st.write(display_be)
            st.dataframe(display_be, use_container_width=True)

            st.subheader("🧠 Автоматски коментар")

            st.info(str(value_map.get("Статус", "")))
            st.write(str(value_map.get("Коментар", "")))

            st.success(
                f"""
            Извештајот опфаќа {months_in_period} месеци.

            Компанијата треба да оствари просечна месечна продажба од
            {value_map.get("Просечен месечен Break-even", 0):,.0f} денари
            за да ги покрие сите трошоци.
            """
                        )
            

        elif izbran_izvestaj == "📄 CFO Snapshot":

            st.title("📄 CFO Snapshot")

            be_cfo = calculate_break_even(
                df=df,
                bu=bu,
                sales_aops=["202"],
                pct_402=70,
                pct_403=70
            )

            be_map = dict(zip(be_cfo["Показател"], be_cfo["Вредност"]))

            sales_revenue = float(be_map.get("Приход од продажба", 0))
            break_even_sales = float(be_map.get("Break-even продажба", 0))
            margin_of_safety = float(be_map.get("Над Break-even", 0))
            safety_margin_pct = float(be_map.get("Безбедносна маргина %", 0))

            net_profit = float(
                bu.loc[bu["AOP"].astype(str).str.zfill(3) == "255", "Iznos"].sum()
            )
            net_loss = float(
                bu.loc[bu["AOP"].astype(str).str.zfill(3) == "256", "Iznos"].sum()
            )
            net_result = net_profit - net_loss

            def get_kpi_raw(kpi_df, naziv_contains):
                try:
                    match = kpi_df[
                        kpi_df["Naziv"].astype(str).str.contains(
                            naziv_contains,
                            case=False,
                            na=False
                        )
                    ]
                    if match.empty:
                        return 0.0
                    return float(match.iloc[0].get("RawValue", 0.0))
                except Exception:
                    return 0.0

            current_ratio = get_kpi_raw(kpi, "ТЕКОВНА ЛИКВИДНОСТ")
            profit_margin = get_kpi_raw(kpi, "ПРОФИТНА МАРЖА")
            debt_ratio = get_kpi_raw(kpi, "ВКУПНА ЗАДОЛЖЕНОСТ")

            score = 0

            if net_result > 0 and profit_margin >= 0.15:
                score += 35
            elif net_result > 0 and profit_margin >= 0.05:
                score += 25
            elif net_result > 0:
                score += 15

            if current_ratio >= 1.5:
                score += 35
            elif current_ratio >= 1.0:
                score += 25
            elif current_ratio >= 0.7:
                score += 15
            else:
                score += 5

            if safety_margin_pct >= 0.30:
                score += 30
            elif safety_margin_pct >= 0.10:
                score += 20
            elif safety_margin_pct > 0:
                score += 10

            if score >= 80:
                cfo_status = "🟢 ДОБРА ФИНАНСИСКА СОСТОЈБА"
                status_color = "#15803d"
            elif score >= 60:
                cfo_status = "🟡 СТАБИЛНА СОСТОЈБА СО ПОТРЕБА ОД СЛЕДЕЊЕ"
                status_color = "#ca8a04"
            else:
                cfo_status = "🔴 ПОТРЕБНО Е ВНИМАНИЕ"
                status_color = "#dc2626"

            st.markdown(
                f"""
                <div style="
                    text-align:center;
                    padding:28px;
                    border-radius:22px;
                    background:linear-gradient(135deg,#f8fafc,#e0f2fe);
                    border:1px solid #dbeafe;
                    margin-bottom:24px;
                ">
                    <h2 style="margin:0;color:#0f172a;">🏆 SIGMA CFO SCORE</h2>
                    <div style="font-size:64px;font-weight:800;color:#1e3a8a;margin:10px 0;">
                        {score}/100
                    </div>
                    <h3 style="color:{status_color};margin:0;">
                        {cfo_status}
                    </h3>
                </div>
                """,
                unsafe_allow_html=True
            )

            summary_lines = []

            if net_result > 0:
                summary_lines.append(
                    f"Компанијата оствари нето добивка од {net_result:,.0f} денари."
                )
            else:
                summary_lines.append(
                    f"Компанијата оствари загуба од {abs(net_result):,.0f} денари."
                )

            if safety_margin_pct >= 0.30:
                summary_lines.append(
                    f"Продажбата е {safety_margin_pct*100:.2f}% над break-even точката што укажува на добра оперативна стабилност."
                )
            elif safety_margin_pct >= 0.10:
                summary_lines.append(
                    f"Продажбата е {safety_margin_pct*100:.2f}% над break-even точката и потребно е редовно следење на трошоците."
                )
            else:
                summary_lines.append(
                    f"Компанијата работи блиску до break-even точката и е чувствителна на пад на продажбата."
                )

            if current_ratio >= 1.5:
                summary_lines.append(
                    f"Ликвидноста е на добро ниво ({current_ratio:.2f})."
                )
            elif current_ratio >= 1:
                summary_lines.append(
                    f"Ликвидноста е прифатлива ({current_ratio:.2f}) но треба да се следи."
                )
            else:
                summary_lines.append(
                    f"Тековната ликвидност од {current_ratio:.2f} е под препорачаното ниво и бара внимание кон наплатата и обртните средства."
                )

            executive_summary = "\n\n".join(summary_lines)




            summary_lines = []

            if net_result > 0:
                summary_lines.append(
                    f"Компанијата оствари нето добивка од {net_result:,.0f} денари."
                )
            else:
                summary_lines.append(
                    f"Компанијата оствари загуба од {abs(net_result):,.0f} денари."
                )

            if safety_margin_pct >= 0.30:
                summary_lines.append(
                    f"Продажбата е {safety_margin_pct*100:.2f}% над break-even точката што укажува на добра оперативна стабилност."
                )
            elif safety_margin_pct >= 0.10:
                summary_lines.append(
                    f"Продажбата е {safety_margin_pct*100:.2f}% над break-even точката и потребно е редовно следење на трошоците."
                )
            else:
                summary_lines.append(
                    f"Компанијата работи блиску до break-even точката и е чувствителна на пад на продажбата."
                )

            if current_ratio >= 1.5:
                summary_lines.append(
                    f"Ликвидноста е на добро ниво ({current_ratio:.2f})."
                )
            elif current_ratio >= 1:
                summary_lines.append(
                    f"Ликвидноста е прифатлива ({current_ratio:.2f}) но треба да се следи."
                )
            else:
                summary_lines.append(
                    f"Тековната ликвидност од {current_ratio:.2f} е под препорачаното ниво и бара внимание кон наплатата и обртните средства."
                )

            executive_summary = "\n\n".join(summary_lines)

            st.subheader("📌 Executive Summary")

            st.success(executive_summary)

            st.subheader("📊 Клучни показатели")

            c1, c2, c3, c4, c5 = st.columns(5)

            c1.metric("Продажба", f"{sales_revenue:,.0f}")
            c2.metric("Нето добивка", f"{net_result:,.0f}")
            c3.metric("EBITDA", "-")
            c4.metric("Break-even", f"{break_even_sales:,.0f}")
            c5.metric("Безбедносна маргина", f"{safety_margin_pct * 100:.2f}%")

            st.subheader("🎯 Break-even Snapshot")

            b1, b2, b3 = st.columns(3)

            b1.metric("Реална продажба", f"{sales_revenue:,.0f}")
            b2.metric("Break-even продажба", f"{break_even_sales:,.0f}")
            b3.metric("Над Break-even", f"{margin_of_safety:,.0f}")

            st.progress(min(max(safety_margin_pct, 0), 1))
            st.caption(f"Безбедносна маргина: {safety_margin_pct * 100:.2f}%")

            st.subheader("🚦 Топ 3 CFO препораки")

            st.markdown(
                """
                🟢 Компанијата работи профитабилно.

                🟢 Оперативниот ризик е низок според Break-even анализата.

                🟡 Следен фокус: анализа на залихи, наплата и обврски.
                """
            )

            st.subheader("💧 Liquidity & Debt Snapshot")

            l1, l2, l3, l4 = st.columns(4)

            l1.metric(
                "Тековна ликвидност",
                f"{current_ratio:.2f}"
            )

            l2.metric(
                "Вкупна задолженост",
                f"{debt_ratio * 100:.2f}%"
            )

            quick_ratio = get_kpi_raw(
                kpi,
                "МОМЕНТАЛНА ЛИКВИДНОСТ"
            )

            l3.metric(
                "Quick Ratio",
                f"{quick_ratio:.2f}"
            )

            financial_stability = get_kpi_raw(
                kpi,
                "ФИНАНСИСКА СТАБИЛНОСТ"
            )

            l4.metric(
                "Финансиска стабилност",
                f"{financial_stability * 100 :.2f}%"
            )

            liquidity_text = []

            if current_ratio < 1:
                liquidity_text.append(
                    "🔴 Тековната ликвидност е под препорачаното ниво."
                )
            else:
                liquidity_text.append(
                    "🟢 Ликвидноста е задоволителна."
                )

            #st.write("Debt ratio =", debt_ratio)
            if debt_ratio > 0.80:
                liquidity_text.append(
                    "🔴 Задолженоста е висока и претставува финансиски ризик."
                )
            elif debt_ratio > 0.60:
                liquidity_text.append(
                    "🟠 Задолженоста е релативно висока и бара внимание."
                )
            elif debt_ratio > 0.40:
                liquidity_text.append(
                    "🟡 Задолженоста е умерена и треба да се следи."
                )

            else:
                liquidity_text.append(
                    "🟢 Задолженоста е на прифатливо ниво."
                )

            st.info("\n\n".join(liquidity_text))

            risk_notes = []

            if current_ratio < 1 and debt_ratio > 60:
                risk_notes.append(
                    "⚠️ Комбинацијата на ниска ликвидност и повисока задолженост претставува најзначаен финансиски ризик во моментот."
                )

            if current_ratio < 1 and safety_margin_pct > 0.30:
                risk_notes.append(
                    "ℹ️ И покрај слабата ликвидност, компанијата има добра оперативна сигурност преку задоволителна безбедносна маргина."
                )

            if net_result > 0 and current_ratio < 1:
                risk_notes.append(
                    "ℹ️ Компанијата е профитабилна, но добивката не се претвора доволно брзо во готовина."
                )

            for note in risk_notes:
                st.warning(note)


        elif izbran_izvestaj == "📘 CFO Извештај":

            st.title("📘 CFO Извештај")
            st.caption("Комбиниран CFO извештај со избор на секции за печатење / зачувување како PDF.")

            if user_plan != "PREMIUM":
                premium_lock("CFO Извештај")
            else:
                col1, col2 = st.columns(2)

                with col1:
                    company_name = st.text_input(
                        "Назив на компанија",
                        value="Клиент / Компанија",
                        key="cfo_report_company_name"
                    )

                with col2:
                    period_label = st.text_input(
                        "Период",
                        value=datetime.now().strftime("%d.%m.%Y"),
                        key="cfo_report_period"
                    )

                st.subheader("Избери кои извештаи да бидат вклучени")

                c1, c2, c3 = st.columns(3)

                with c1:
                    include_snapshot = st.checkbox("📄 CFO Snapshot", value=True, key="cfo_include_snapshot")
                    include_bu = st.checkbox("📈 Биланс на успех", value=True, key="cfo_include_bu")

                with c2:
                    include_bs = st.checkbox("🏦 Биланс на состојба", value=True, key="cfo_include_bs")
                    include_cf = st.checkbox("💧 Cash Flow", value=True, key="cfo_include_cf")

                with c3:
                    include_kpi = st.checkbox("📊 KPI Dashboard", value=True, key="cfo_include_kpi")
                    include_sd = st.checkbox("🧾 Сеопфатна добивка", value=True, key="cfo_include_sd")

                if st.button("📄 Генерирај CFO извештај", type="primary", key="generate_cfo_report_btn"):

                    sections_html = []

                    if include_snapshot:
                        cfo_snapshot_df = build_cfo_snapshot_df_for_report(df, bu, kpi)
                        sections_html.append(
                            cfo_df_to_html(
                                cfo_snapshot_df,
                                "CFO Snapshot"
                            )
                        )

                    if include_bu:
                        bu_cfo = cfo_remove_empty_report_rows(
                            bu,
                            value_columns=["Iznos"]
                        )

                        sections_html.append(
                            cfo_df_to_html(
                                bu_cfo,
                                "Биланс на успех",
                                columns=["AOP", "Pozicija", "Iznos"]
                            )
                        )

                    if include_bs:
                        bs_cfo = cfo_remove_empty_report_rows(
                            bs,
                            value_columns=["Tekovna godina", "Prethodna godina"]
                        )

                        sections_html.append(
                            cfo_df_to_html(
                                bs_cfo,
                                "Биланс на состојба",
                                columns=["AOP", "Pozicija", "Tekovna godina", "Prethodna godina"]
                            )
                        )

                    if include_cf:
                        if cf is not None:
                            cf_cfo = cfo_remove_empty_report_rows(
                                cf,
                                value_columns=["Iznos"]
                            )
                            sections_html.append(
                                cfo_df_to_html(
                                    cf_cfo,
                                    "Cash Flow"
                                )
                            )
                        else:
                            sections_html.append(
                                """
                                <section class="report-section page-break">
                                    <h2>Cash Flow</h2>
                                    <p class="muted">Cash Flow не е вчитан или не е пресметан.</p>
                                </section>
                                """
                            )

                    if include_kpi:
                        if kpi is not None:
                            sections_html.append(
                                cfo_df_to_html(
                                    kpi,
                                    "KPI Dashboard",
                                    columns=["Kategorija", "Naziv", "Vrednost", "Status"]
                                )
                            )
                        else:
                            sections_html.append(
                                """
                                <section class="report-section page-break">
                                    <h2>KPI Dashboard</h2>
                                    <p class="muted">KPI не е вчитан или не е пресметан.</p>
                                </section>
                                """
                            )

                    if include_sd:
                        try:
                            sd_report = sd[["Red", "Naziv", "Tekovna godina"]].copy()
                        except Exception:
                            sd_report = sd

                        sd_cfo = cfo_remove_empty_report_rows(
                            sd_report,
                            value_columns=["Tekovna godina"]
                        )

                        sections_html.append(
                            cfo_df_to_html(
                                sd_cfo,
                                "Сеопфатна добивка"
                            )
                        )

                    if not sections_html:
                        st.warning("Нема избрано ниту еден извештај.")
                    else:
                        cfo_html = build_cfo_report_html(
                            company_name=company_name,
                            period_label=period_label,
                            sections_html=sections_html
                        )

                        st.success("CFO извештајот е генериран.")

                        st.download_button(
                            "⬇️ Преземи CFO извештај како HTML / Print to PDF",
                            data=cfo_html.encode("utf-8"),
                            file_name="CFO_izvestaj.html",
                            mime="text/html",
                            key="download_cfo_report_html"
                        )

                        with st.expander("👁️ Преглед на CFO извештај", expanded=True):
                            components.html(cfo_html, height=900, scrolling=True)


        elif izbran_izvestaj == "💎 Valuation Dashboard":
            st.subheader("📈 EBITDA Valuation / ПРОЦЕНКА")

            if user_plan != "PREMIUM":

                st.warning("🔒 Valuation Dashboard е PREMIUM модул. Бројките се скриени.")

                col1, col2, col3 = st.columns(3)

                col1.metric("EBITDA / ЕБИТДА", "••••••")
                col2.metric("Enterprise Value / ВРЕДНОСТ НА КОМПАНИЈАТА", "••••••")
                col3.metric("Equity Value / ВРЕДНОСТ НА КАПИТАЛОТ", "••••••")

                col4, col5 = st.columns(2)

                col4.metric("Cash / Парични средства", "••••••")
                col5.metric("Debt / Финансиски долг", "••••••")

                st.subheader("📊 Valuation Range / Опсег на проценка")

                c1, c2, c3 = st.columns(3)

                c1.metric("Conservative / Конзервативна (3x)", "••••••")
                c2.metric("Fair Value / Реална вредност (4x)", "••••••")
                c3.metric("Optimistic / Оптимистичка вредност (5x)", "••••••")

                premium_lock("Valuation Dashboard")

            else:

                multiple = st.number_input(
                    "EBITDA Multiple",
                    min_value=1.0,
                    max_value=10.0,
                    value=4.0,
                    step=0.5
                )

                enterprise_value = ebitda * multiple
                equity_value = enterprise_value + cash - debt

                conservative_value = ebitda * 3
                fair_value = ebitda * 4
                optimistic_value = ebitda * 5

                conservative_equity = conservative_value + cash - debt
                fair_equity = fair_value + cash - debt
                optimistic_equity = optimistic_value + cash - debt

                col1, col2, col3 = st.columns(3)

                col1.metric("EBITDA / ЕБИТДА", f"{ebitda:,.0f}")
                col2.metric("Enterprise Value / ВРЕДНОСТ НА КОМПАНИЈАТА", f"{enterprise_value:,.0f}")
                col3.metric("Equity Value / ВРЕДНОСТ НА КАПИТАЛОТ", f"{equity_value:,.0f}")

                col4, col5 = st.columns(2)

                col4.metric("Cash / Парични средства", f"{cash:,.0f}")
                col5.metric("Debt / Финансиски долг", f"{debt:,.0f}")

                st.subheader("📊 Valuation Range / Опсег на проценка")

                c1, c2, c3 = st.columns(3)

                c1.metric("Conservative / Конзервативна (3x)", f"{conservative_equity:,.0f}")
                c2.metric("Fair Value / Реална вредност (4x)", f"{fair_equity:,.0f}")
                c3.metric("Optimistic / Оптимистичка вредност (5x)", f"{optimistic_equity:,.0f}")

                st.subheader("📉 DCF Valuation / Проценка по DCF метод")

                dcf_years = st.number_input(
                    "Период на проекција - години",
                    min_value=1,
                    max_value=10,
                    value=5,
                    step=1
                )

                growth_rate = st.number_input(
                    "Годишен раст на Cash Flow (%)",
                    min_value=-20.0,
                    max_value=50.0,
                    value=5.0,
                    step=0.5
                ) / 100

                discount_rate = st.number_input(
                    "Discount rate / Стапка на дисконтирање (%)",
                    min_value=1.0,
                    max_value=50.0,
                    value=12.0,
                    step=0.5
                ) / 100

                terminal_growth = st.number_input(
                    "Terminal growth rate (%)",
                    min_value=0.0,
                    max_value=10.0,
                    value=2.0,
                    step=0.5
                ) / 100

                base_cash_flow = ebitda

                dcf_value = 0

                for year in range(1, dcf_years + 1):
                    projected_cf = base_cash_flow * ((1 + growth_rate) ** year)
                    discounted_cf = projected_cf / ((1 + discount_rate) ** year)
                    dcf_value += discounted_cf

                terminal_value = (
                    projected_cf * (1 + terminal_growth)
                ) / (discount_rate - terminal_growth)

                discounted_terminal_value = terminal_value / ((1 + discount_rate) ** dcf_years)

                enterprise_value_dcf = dcf_value + discounted_terminal_value
                equity_value_dcf = enterprise_value_dcf + cash - debt

                d1, d2, d3 = st.columns(3)

                d1.metric("DCF Enterprise Value", f"{enterprise_value_dcf:,.0f}")
                d2.metric("DCF Equity Value", f"{equity_value_dcf:,.0f}")
                d3.metric("Terminal Value", f"{discounted_terminal_value:,.0f}")

        export_sheets = {
            "Zaklucen list": df_total,
            "Zbir po klasi": class_summary if "class_summary" in locals() else build_konto_class_summary(df),
            "Zbir po klasi rules": class_summary_rules if "class_summary_rules" in locals() else build_konto_class_summary(df, mode="rules"),
            "Bilans uspeh AOP": bu,
            "Bilans sostojba AOP": bs,
            "BS diagnostika": bs_diagnostic if "bs_diagnostic" in locals() else build_bs_class_diagnostic(df, bu, bs),
        }
        try:
            export_sheets["BU rule trace"] = bu_rule_trace if "bu_rule_trace" in locals() else build_rule_trace(df, pravila_file, "Pravila Bilans Uspeh")
            export_sheets["BS rule trace"] = bs_rule_trace if "bs_rule_trace" in locals() else build_rule_trace(df, pravila_file, "Pravila bilans Sostojba", year_mode="current")
        except Exception:
            pass
        if cf is not None:
            export_sheets["Cash Flow"] = cf
        if kpi is not None:
            export_sheets["KPI"] = kpi
        excel_data = export_excel(export_sheets)
        st.download_button("📥 Export Excel", data=excel_data, file_name="aop_finansiski_izvestai.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="export_excel_btn")

        pdf_bu = export_single_pdf(bu, "Bilans na uspeh", "bilans_na_uspeh")
        st.download_button(
    "📄 Export PDF - БИЛАНС НА УСПЕХ",
    data=pdf_bu,
    file_name="bilans_na_uspeh.pdf",
    mime="application/pdf",
    key="pdf_bu_btn"
)

        pdf_bs = export_single_pdf(bs, "Bilans na sostojba", "bilans_na_sostojba")
        st.download_button(
             "📄 Export PDF - БИЛАНС НА СОСТОЈБА",
             data=pdf_bs,
             file_name="bilans_na_sostojba.pdf",
             mime="application/pdf",
             key="pdf_bs_btn"
)

        pdf_sd = export_single_pdf(
             sd[["Red", "Naziv", "Tekovna godina"]],
             "Seopfatna dobivka",
             "seopfatna_dobivka"
)

        st.download_button(
             "📄 Export PDF - СЕОПФАТНА ДОБИВКА",
             data=pdf_sd,
              file_name="seopfatna_dobivka.pdf",
              mime="application/pdf",
              key="pdf_sd_btn"
)

        if cf is not None:
             pdf_cf = export_cash_flow_pdf(cf)

             st.download_button(
             "📄 Export PDF - Cash Flow",
             data=pdf_cf,
             file_name="cash_flow.pdf",
             mime="application/pdf",
              key="pdf_cf_btn"
    )

    except Exception as e:
             import traceback
             st.code(traceback.format_exc())   